"""Módulo separado para la exportación de etiquetas a Excel.

Este archivo existe como módulo independiente porque su función es larga y
contiene caracteres que dificultan el edit del app.py principal (que tiene
line endings CRLF de Windows). Al mantenerlo aparte, evitamos truncar
app.py durante ediciones.

Registro: llamar a `registrar_rutas_etiquetas(app, db, Item, Especialidad,
                                             registrar_auditoria,
                                             login_requerido)` en app.py.
"""
from flask import request, session, flash, redirect, url_for, send_file
from datetime import datetime
import io
import os
import shutil
import tempfile


def registrar_rutas_etiquetas(app, db, Item, Especialidad,
                              registrar_auditoria, login_requerido):
    """Registra la ruta /exportar_etiquetas_excel en la app Flask."""

    @app.route('/exportar_etiquetas_excel')
    @login_requerido
    def exportar_etiquetas_excel():
        """Genera un archivo XLSX con todas las etiquetas del inventario de
        la especialidad activa. Diseñado para imprimir en A4 portrait:
        3 etiquetas por fila, con código de barras Code128 como imagen.
        """
        # Resolución de especialidad
        if session.get('usuario_rol') == 'Admin':
            esp_id = (request.args.get('especialidad_id', type=int)
                      or session.get('usuario_especialidad_id'))
        else:
            esp_id = session.get('usuario_especialidad_id')
        if not esp_id:
            flash("Sin especialidad activa.")
            return redirect(url_for('ver_inventario'))

        esp = Especialidad.query.get(esp_id)
        items = Item.query.filter_by(especialidad_id=esp_id).order_by(
            Item.categoria.asc(), Item.nombre.asc()).all()
        if not items:
            flash("No hay ítems para exportar.")
            return redirect(url_for('ver_inventario'))

        # Imports pesados solo aquí
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.drawing.image import Image as XLImage
        import barcode
        from barcode.writer import ImageWriter

        wb = Workbook()
        ws = wb.active
        ws.title = "Etiquetas"

        # Estilos
        NAVY = "0F2C5C"; GOLD = "B45309"; GREY = "94A3B8"
        thin = Side(border_style="thin", color="CBD5E1")
        border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
        font_titulo = Font(name="Arial", size=14, bold=True, color=NAVY)
        font_sub = Font(name="Arial", size=10, italic=True, color=GREY)
        font_nombre = Font(name="Arial", size=10, bold=True, color=NAVY)
        font_meta = Font(name="Arial", size=8, color="475569")
        font_codigo = Font(name="Consolas", size=9, color=GOLD, bold=True)

        # Encabezado del documento
        ws.merge_cells("A1:C1")
        c = ws["A1"]
        c.value = f"Etiquetas de Inventario - {esp.nombre if esp else 'Especialidad'}"
        c.font = font_titulo
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        ws.merge_cells("A2:C2")
        c = ws["A2"]
        c.value = (f"Generado el {datetime.now().strftime('%d-%m-%Y %H:%M')}  |  "
                   f"{len(items)} items  |  PanolERP")
        c.font = font_sub
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 16

        # Layout: 3 etiquetas por fila
        ETIQUETAS_POR_FILA = 3
        col_letters = ["A", "B", "C"]
        for letra in col_letters:
            ws.column_dimensions[letra].width = 34

        tmp_dir = tempfile.mkdtemp(prefix="etiquetas_")

        fila_actual = 4
        for idx, item in enumerate(items):
            pos = idx % ETIQUETAS_POR_FILA
            col = col_letters[pos]

            if pos == 0:
                ws.row_dimensions[fila_actual].height = 90
                ws.row_dimensions[fila_actual + 1].height = 18
                ws.row_dimensions[fila_actual + 2].height = 15
                ws.row_dimensions[fila_actual + 3].height = 15
                ws.row_dimensions[fila_actual + 4].height = 10

            codigo = (item.codigo_barras or f"ITEM{item.id:06d}").strip()
            img_path = os.path.join(tmp_dir, f"{item.id}")
            try:
                Code128 = barcode.get_barcode_class('code128')
                bar = Code128(codigo, writer=ImageWriter())
                bar.save(img_path, options={
                    'module_height': 12.0,
                    'font_size': 8,
                    'text_distance': 3.0,
                    'quiet_zone': 2.0,
                    'write_text': False,
                })
                img_full = img_path + '.png'
            except Exception:
                img_full = None

            celda_img = f"{col}{fila_actual}"
            if img_full and os.path.exists(img_full):
                xlimg = XLImage(img_full)
                xlimg.width = 220
                xlimg.height = 80
                ws.add_image(xlimg, celda_img)
            ws[celda_img].border = border_all

            celda_codigo = f"{col}{fila_actual + 1}"
            c = ws[celda_codigo]
            c.value = codigo
            c.font = font_codigo
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border_all

            celda_nombre = f"{col}{fila_actual + 2}"
            c = ws[celda_nombre]
            c.value = item.nombre[:60]
            c.font = font_nombre
            c.alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)
            c.border = border_all

            celda_meta = f"{col}{fila_actual + 3}"
            c = ws[celda_meta]
            partes = []
            if item.marca: partes.append(item.marca)
            if item.modelo: partes.append(item.modelo)
            if item.ubicacion: partes.append(f"Ubic: {item.ubicacion}")
            c.value = "  |  ".join(partes) if partes else (item.categoria or "")
            c.font = font_meta
            c.alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)
            c.border = border_all

            if pos == ETIQUETAS_POR_FILA - 1:
                fila_actual += 5

        # Configuración de impresión
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        try:
            shutil.rmtree(tmp_dir)
        except Exception:
            pass

        registrar_auditoria('exportar', 'Item', 0,
                            valores_nuevos={
                                'formato': 'xlsx_etiquetas',
                                'especialidad_id': esp_id,
                                'items_incluidos': len(items),
                            })

        nombre_arch = (f"Etiquetas_"
                       f"{(esp.nombre if esp else 'inventario').replace(' ', '_')}"
                       f"_{datetime.now().strftime('%Y%m%d')}.xlsx")
        return send_file(
            output,
            as_attachment=True,
            download_name=nombre_arch,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
