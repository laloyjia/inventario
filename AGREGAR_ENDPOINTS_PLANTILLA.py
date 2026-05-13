# -*- coding: utf-8 -*-
"""
AGREGAR_ENDPOINTS_PLANTILLA.py
==============================
Inserta TODOS los endpoints faltantes en app.py de una sola pasada.

Endpoints que agrega (si no existen):
  - /descargar_plantilla
  - /descargar_plantilla_alumnos
  - /exportar_inventario (alias de /exportar_excel)
  - /panoleros_dia/agregar
  - /panoleros_dia/quitar/<int:pd_id>
  - /panoleros_dia/limpiar
  - /regenerar_codigo_alumno/<int:est_id>  (si no estaba)
  - /admin/reporte_mermas
  - /admin/exportar_completo
  - /admin/buscar
  - /admin/cambiar_password

Backup automatico antes de tocar nada.
Idempotente: si un endpoint ya existe, lo saltea.
Uso:
    python AGREGAR_ENDPOINTS_PLANTILLA.py
"""
import os
import sys
import shutil
import py_compile
from datetime import datetime

AQUI = os.path.dirname(os.path.abspath(__file__))
APP_PY = os.path.join(AQUI, 'app.py')

if not os.path.exists(APP_PY):
    print(f"[ERROR] No encuentro {APP_PY}")
    sys.exit(1)

# Backup
stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
backup = APP_PY + f".bak_endpoints_{stamp}"
shutil.copy2(APP_PY, backup)
print(f"[OK] Backup creado: {os.path.basename(backup)}")

with open(APP_PY, 'r', encoding='utf-8') as f:
    contenido = f.read()

ENDPOINTS = {
    "@app.route('/descargar_plantilla')": '''
@app.route('/descargar_plantilla')
@login_requerido
def descargar_plantilla():
    """Descarga inventario_muestra.xlsx (con fallback a generar al vuelo)."""
    aqui = os.path.dirname(os.path.abspath(__file__))
    plantilla = os.path.join(aqui, 'inventario_muestra.xlsx')
    if os.path.exists(plantilla):
        return send_file(plantilla, as_attachment=True,
                         download_name='plantilla_inventario.xlsx')
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = 'Inventario'
    for c, h in enumerate(['Codigo de barra','Nombre','Descripcion','Categoria',
                           'Cantidad','Ubicacion','Fecha adquisicion','Desgaste ($)',
                           'Costo unitario ($)','Costo total ($)','Imagen de referencia'], 1):
        ws.cell(row=1, column=c, value=h)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='plantilla_inventario.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
''',

    "@app.route('/descargar_plantilla_alumnos')": '''
@app.route('/descargar_plantilla_alumnos')
@login_requerido
def descargar_plantilla_alumnos():
    """Descarga alumnos_muestra.xlsx (con fallback)."""
    aqui = os.path.dirname(os.path.abspath(__file__))
    plantilla = os.path.join(aqui, 'alumnos_muestra.xlsx')
    if os.path.exists(plantilla):
        return send_file(plantilla, as_attachment=True,
                         download_name='plantilla_alumnos.xlsx')
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = 'Alumnos'
    ws.cell(row=1, column=1, value='N de lista')
    ws.cell(row=1, column=2, value='Nombre completo')
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='plantilla_alumnos.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
''',

    "@app.route('/exportar_inventario')": '''
@app.route('/exportar_inventario')
@login_requerido
def exportar_inventario():
    """Alias de /exportar_excel para mantener compatibilidad con la plantilla."""
    return exportar_excel()
''',

    "@app.route('/panoleros_dia/agregar'": '''
@app.route('/panoleros_dia/agregar', methods=['POST'])
@login_requerido
@panolero_o_admin
def panoleros_dia_agregar():
    """Designa un estudiante como panolero del dia. Maximo MAX_PANOLEROS_DIA por especialidad."""
    especialidad_id = session.get('usuario_especialidad_id')
    if not especialidad_id:
        flash("Cuenta sin especialidad.")
        return redirect(url_for('ver_inventario'))
    estudiante_id = request.form.get('estudiante_id', type=int)
    if not estudiante_id:
        flash("Selecciona un estudiante.")
        return redirect(url_for('ver_inventario'))
    activos = PanoleroDesignado.query.filter_by(
        especialidad_id=especialidad_id, activo=True).count()
    if activos >= MAX_PANOLEROS_DIA:
        flash(f"Ya hay {MAX_PANOLEROS_DIA} panoleros del dia activos. Quita alguno antes.")
        return redirect(url_for('ver_inventario'))
    existe = PanoleroDesignado.query.filter_by(
        estudiante_id=estudiante_id, especialidad_id=especialidad_id, activo=True).first()
    if existe:
        flash("Ese estudiante ya esta designado.")
        return redirect(url_for('ver_inventario'))
    nuevo = PanoleroDesignado(
        estudiante_id=estudiante_id,
        especialidad_id=especialidad_id,
        designado_por_id=session.get('usuario_id'),
        activo=True)
    db.session.add(nuevo)
    db.session.commit()
    flash("Panolero del dia agregado.")
    return redirect(url_for('ver_inventario'))
''',

    "@app.route('/panoleros_dia/quitar": '''
@app.route('/panoleros_dia/quitar/<int:pd_id>', methods=['POST'])
@login_requerido
@panolero_o_admin
def panoleros_dia_quitar(pd_id):
    """Da de baja a un panolero del dia."""
    pd = PanoleroDesignado.query.get_or_404(pd_id)
    if session.get('usuario_rol') != 'Admin' and pd.especialidad_id != session.get('usuario_especialidad_id'):
        flash("Sin permiso.")
        return redirect(url_for('ver_inventario'))
    pd.activo = False
    pd.fecha_baja = datetime.utcnow()
    db.session.commit()
    flash("Panolero del dia quitado.")
    return redirect(url_for('ver_inventario'))
''',

    "@app.route('/panoleros_dia/limpiar'": '''
@app.route('/panoleros_dia/limpiar', methods=['POST'])
@login_requerido
@panolero_o_admin
def panoleros_dia_limpiar():
    """Da de baja a TODOS los panoleros del dia de la especialidad."""
    especialidad_id = session.get('usuario_especialidad_id')
    if not especialidad_id:
        flash("Cuenta sin especialidad.")
        return redirect(url_for('ver_inventario'))
    n = 0
    for pd in PanoleroDesignado.query.filter_by(
            especialidad_id=especialidad_id, activo=True).all():
        pd.activo = False
        pd.fecha_baja = datetime.utcnow()
        n += 1
    db.session.commit()
    flash(f"{n} panolero(s) del dia removido(s).")
    return redirect(url_for('ver_inventario'))
''',

    "@app.route('/regenerar_codigo_alumno": '''
@app.route('/regenerar_codigo_alumno/<int:est_id>', methods=['POST'])
@login_requerido
@panolero_o_admin
def regenerar_codigo_alumno(est_id):
    """Genera un nuevo codigo de barras para el alumno."""
    est = Estudiante.query.get_or_404(est_id)
    if session.get('usuario_rol') != 'Admin' and est.especialidad_id != session.get('usuario_especialidad_id'):
        flash("Sin permiso.")
        return redirect(url_for('ver_inventario'))
    est.codigo_barras = generar_codigo_barras_alumno()
    db.session.commit()
    flash(f"Nuevo codigo para {est.nombre}: {est.codigo_barras}")
    return redirect(url_for('ver_inventario'))
''',

    "@app.route('/admin/reporte_mermas')": '''
@app.route('/admin/reporte_mermas')
@login_requerido
@admin_requerido
def admin_reporte_mermas():
    """Placeholder."""
    flash("Reporte de mermas: funcion en construccion.")
    return redirect(url_for('ver_inventario'))
''',

    "@app.route('/admin/exportar_completo')": '''
@app.route('/admin/exportar_completo')
@login_requerido
@admin_requerido
def admin_exportar_completo():
    """Exporta inventario consolidado."""
    items = Item.query.order_by(Item.especialidad_id.asc(),
                                Item.categoria.asc(), Item.nombre.asc()).all()
    df = pd.DataFrame([{
        'Especialidad':   i.especialidad.nombre if i.especialidad else '',
        'Codigo':         i.codigo_barras,
        'Nombre':         i.nombre,
        'Categoria':      i.categoria,
        'Cantidad total': i.cantidad_total,
        'Disponible':     i.cantidad_disponible,
        'Mermada':        i.cantidad_mermada,
        'Ubicacion':      i.ubicacion,
        'Costo unitario': i.precio_unitario or 0,
        'Desgaste':       i.desgaste if hasattr(i, 'desgaste') else 0,
        'Costo total':    (i.precio_unitario or 0) * (i.cantidad_total or 0),
    } for i in items])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Consolidado', index=False)
    buf.seek(0)
    fname = f"inventario_consolidado_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
''',

    "@app.route('/admin/buscar')": '''
@app.route('/admin/buscar')
@login_requerido
@admin_requerido
def admin_buscar():
    """Busqueda global de items para admin."""
    q = (request.args.get('q') or '').strip()
    if not q:
        return render_template('admin_buscar.html', items=[], q='')
    like = f"%{q}%"
    items = Item.query.filter(
        db.or_(
            Item.nombre.ilike(like),
            Item.codigo_barras.ilike(like),
            Item.autor.ilike(like) if hasattr(Item, 'autor') else False,
            Item.isbn.ilike(like) if hasattr(Item, 'isbn') else False,
        )
    ).order_by(Item.especialidad_id.asc(), Item.nombre.asc()).limit(200).all()
    return render_template('admin_buscar.html', items=items, q=q)
''',

    "@app.route('/admin/cambiar_password'": '''
@app.route('/admin/cambiar_password', methods=['GET', 'POST'])
@login_requerido
def admin_cambiar_password():
    """Cambiar contrasena del usuario logueado."""
    user = Usuario.query.get(session.get('usuario_id'))
    if not user:
        flash("Sesion invalida.")
        return redirect(url_for('login'))
    if request.method == 'POST':
        actual = request.form.get('password_actual', '')
        nueva = request.form.get('password_nueva', '')
        confirm = request.form.get('password_confirm', '')
        if not check_password_hash(user.password_hash, actual):
            flash("Contrasena actual incorrecta.")
            return redirect(url_for('admin_cambiar_password'))
        if len(nueva) < 8:
            flash("La nueva contrasena debe tener al menos 8 caracteres.")
            return redirect(url_for('admin_cambiar_password'))
        if nueva != confirm:
            flash("Las contrasenas no coinciden.")
            return redirect(url_for('admin_cambiar_password'))
        user.password_hash = generate_password_hash(nueva)
        user.must_change_password = False
        db.session.commit()
        flash("Contrasena actualizada.")
        return redirect(url_for('index'))
    return render_template('cambiar_password.html', usuario=user)
''',
}


def encontrar_punto_insercion(contenido):
    """Devuelve el indice donde insertar (justo antes de 'if __name__')."""
    idx = contenido.rfind("if __name__ == '__main__':")
    if idx == -1:
        return None
    return idx


def main():
    global contenido
    agregados = []
    omitidos = []
    bloque_a_insertar = ""

    for clave, codigo in ENDPOINTS.items():
        # Buscar si ya existe (clave es un fragmento unico del @app.route)
        if clave in contenido:
            omitidos.append(clave)
            continue
        bloque_a_insertar += codigo + "\n"
        agregados.append(clave)

    if not bloque_a_insertar:
        print("[OK] Todos los endpoints ya estan presentes. Nada que hacer.")
        return

    idx = encontrar_punto_insercion(contenido)
    if idx is None:
        print("[ERROR] No encuentro 'if __name__' en app.py. Esta truncado?")
        print("        Restaura con: git checkout -- app.py")
        sys.exit(2)

    # Tambien necesitamos que existan los decoradores @panolero_o_admin y @admin_requerido
    # y que existan los imports. Verificar.
    decorador = 'panolero_o_admin'
    if decorador not in contenido and 'pañolero_o_admin' in contenido:
        # En el codigo real se uso "panolero_o_admin" (con n con tilde). Reemplazamos.
        for k in list(ENDPOINTS.keys()):
            ENDPOINTS[k] = ENDPOINTS[k].replace('panolero_o_admin', 'pañolero_o_admin')
        bloque_a_insertar = bloque_a_insertar.replace('panolero_o_admin', 'pañolero_o_admin')

    nuevo = contenido[:idx] + bloque_a_insertar + "\n\n" + contenido[idx:]

    with open(APP_PY, 'w', encoding='utf-8') as f:
        f.write(nuevo)

    # Verificar sintaxis
    try:
        py_compile.compile(APP_PY, doraise=True)
    except py_compile.PyCompileError as e:
        print(f"[ERROR] El archivo modificado tiene errores: {e}")
        shutil.copy2(backup, APP_PY)
        print("[OK] Restaurado backup.")
        sys.exit(3)

    print(f"[OK] {len(agregados)} endpoints agregados:")
    for a in agregados:
        print(f"     + {a}")
    if omitidos:
        print(f"[INFO] {len(omitidos)} ya existian (omitidos):")
        for o in omitidos:
            print(f"     . {o}")
    print()
    print("Siguientes pasos:")
    print("  wc -l app.py    # debe haber crecido")
    print("  tail -1 app.py  # debe terminar con app.run")
    print("  git add app.py AGREGAR_ENDPOINTS_PLANTILLA.py")
    print("  git commit -m 'fix: reagregar 11 endpoints perdidos en truncamientos'")
    print("  git push")


if __name__ == '__main__':
    main()
