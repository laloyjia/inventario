from flask import Flask, request, session, redirect, url_for, flash, render_template, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime, timedelta
import json
import pandas as pd
import io
import os
import traceback

app = Flask(__name__)

# ============================================================
# CONFIGURACIÓN — soporta tres modos de despliegue:
#   1. Local desarrollo:   SQLite en instance/inventario.db
#   2. Nodo en red local:  SQLite local + sincronización a un admin
#   3. Cloud (Render/Railway/etc): PostgreSQL via DATABASE_URL
# ============================================================

app.secret_key = os.getenv('PANOL_SECRET_KEY', 'llave_super_secreta_enterprise_v5_multiespecialidad')

# ────────────────────────────────────────────────────────────────────
# Handler global de errores 500: escribe el traceback COMPLETO al
# stdout (que Render captura en su vista de Logs) y devuelve un
# mensaje con la primera línea del error a la interfaz. Esto elimina
# las pantallas "Error Interno" en blanco que no dicen nada.
# ────────────────────────────────────────────────────────────────────
@app.errorhandler(500)
def _err500(e):
    tb = traceback.format_exc()
    print("═══ ERROR 500 ═══")
    print(f"Ruta: {request.path}  Método: {request.method}")
    print(f"Usuario: {session.get('usuario_id')} / rol: {session.get('usuario_rol')}")
    print(tb)
    print("═══ FIN ERROR 500 ═══")
    linea = tb.strip().splitlines()[-1] if tb else "sin detalle"
    return (
        "<h1>Error Interno del Servidor</h1>"
        "<p>Detalle técnico (para reportar al administrador):</p>"
        f"<pre style='background:#fee;padding:10px;border-left:4px solid #b91c1c;'>{linea}</pre>"
        "<p><a href='/'>Volver al inicio</a></p>"
    ), 500

@app.errorhandler(Exception)
def _err_generico(e):
    # Cualquier excepción no-HTTP se re-lanza para que Flask la trate como 500,
    # pero también la logueamos por si es un tipo específico (IntegrityError, etc.)
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return e
    tb = traceback.format_exc()
    print("═══ EXCEPCIÓN NO CAPTURADA ═══")
    print(f"Ruta: {request.path}  Método: {request.method}")
    print(f"Tipo: {type(e).__name__}")
    print(tb)
    print("═══ FIN EXCEPCIÓN ═══")
    linea = f"{type(e).__name__}: {str(e)[:200]}"
    return (
        "<h1>Error Interno del Servidor</h1>"
        "<p>Detalle técnico (para reportar al administrador):</p>"
        f"<pre style='background:#fee;padding:10px;border-left:4px solid #b91c1c;'>{linea}</pre>"
        "<p><a href='/'>Volver al inicio</a></p>"
    ), 500

# Carpeta instance para SQLite local
db_folder = os.path.join(os.path.dirname(__file__), 'instance')
os.makedirs(db_folder, exist_ok=True)

# Identificación del nodo
NODO_ID = os.getenv('PANOL_NODO', 'local')
PANOL_ADMIN_URL = os.getenv('PANOL_ADMIN_URL', '').rstrip('/')
ES_ADMIN_CENTRAL = (PANOL_ADMIN_URL == '')
PANOL_SYNC_TOKEN = os.getenv('PANOL_SYNC_TOKEN', 'CAMBIAR_ESTE_TOKEN_PRODUCCION')

# === Conexión a la base de datos ===
DATABASE_URL = os.getenv('DATABASE_URL', '').strip()
if DATABASE_URL:
    # Render entrega URLs como "postgres://..." pero SQLAlchemy 2.x exige "postgresql://"
    if DATABASE_URL.startswith('postgres://'):
        DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
    USANDO_POSTGRES = DATABASE_URL.startswith('postgresql')
    print(f"[DB] Usando {'PostgreSQL' if USANDO_POSTGRES else 'externa'}: {DATABASE_URL.split('@')[-1] if '@' in DATABASE_URL else DATABASE_URL}")
else:
    # Fallback: SQLite local
    DB_PATH = os.getenv('PANOL_DB_PATH', os.path.join(db_folder, 'inventario.db'))
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
    USANDO_POSTGRES = False
    print(f"[DB] Usando SQLite local: {DB_PATH}")

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,        # detecta conexiones muertas (importante en cloud)
    'pool_recycle': 280,          # recicla conexiones cada ~5 min
}

# === Seguridad de cookies (importante en cloud) ===
EN_PRODUCCION = os.getenv('FLASK_ENV', '').lower() == 'production' or USANDO_POSTGRES
app.config['SESSION_COOKIE_HTTPONLY'] = True       # cookies no accesibles desde JS
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'      # mitiga CSRF cross-site
app.config['SESSION_COOKIE_SECURE'] = EN_PRODUCCION  # solo HTTPS en producción
app.config['PERMANENT_SESSION_LIFETIME'] = 60 * 60 * 8  # 8 horas

# === HTTPS forzado + headers de seguridad (Talisman, opcional) ===
if EN_PRODUCCION and os.getenv('FORCE_HTTPS', 'true').lower() != 'false':
    try:
        from flask_talisman import Talisman
        Talisman(app,
                 force_https=True,
                 strict_transport_security=True,
                 session_cookie_secure=True,
                 content_security_policy={
                     'default-src': ["'self'"],
                     'script-src': ["'self'", "'unsafe-inline'",
                                    'cdn.jsdelivr.net', 'cdnjs.cloudflare.com'],
                     'style-src': ["'self'", "'unsafe-inline'",
                                   'cdnjs.cloudflare.com', 'fonts.googleapis.com'],
                     'font-src': ["'self'", 'cdnjs.cloudflare.com', 'fonts.gstatic.com'],
                     'img-src': ["'self'", 'data:', 'https:'],
                 })
        print("[SEC] Talisman activo: HTTPS forzado + CSP")
    except ImportError:
        print("[SEC] Flask-Talisman no instalado, saltando endurecimiento HTTPS")

# === Rate limiting en login (anti brute-force, opcional) ===
try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    limiter = Limiter(get_remote_address, app=app,
                      default_limits=["500 per hour"],
                      storage_uri="memory://")
    print("[SEC] Rate limiter activo")
except ImportError:
    limiter = None
    print("[SEC] Flask-Limiter no instalado, saltando rate limiting")

db = SQLAlchemy(app)


# ========== FILTROS JINJA ==========

@app.template_filter('clp')
def format_clp(value):
    """Formatea un número como pesos chilenos: 12500 → '$ 12.500'.
    Acepta None, '' o no numérico y devuelve '$ 0'."""
    try:
        n = float(value or 0)
    except (TypeError, ValueError):
        n = 0
    # Separador de miles con punto, sin decimales (CLP no usa centavos)
    entero = int(round(n))
    return "$ " + f"{entero:,}".replace(",", ".")


# ========== TABLAS DE BASE DE DATOS (PanolERP fase 2) ==========

class Especialidad(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), unique=True, nullable=False)
    descripcion = db.Column(db.Text)
    color = db.Column(db.String(7), default="#2563eb")
    activa = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    # Tipo de área: PANOL_TP, BIBLIOTECA, DEPORTIVO, INFORMATICA, GENERAL
    # Define qué campos del Item se muestran/usan en el formulario y la lista.
    tipo_area = db.Column(db.String(20), default='GENERAL', nullable=False,
                          server_default='GENERAL')
    usuarios = db.relationship('Usuario', backref='especialidad_asignada', lazy=True)
    items = db.relationship('Item', backref='especialidad', lazy=True)

class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    username = db.Column(db.String(50), unique=True, nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=True)
    password_hash = db.Column(db.String(200), nullable=False)
    rol = db.Column(db.String(50), default='Pañolero')
    especialidad_id = db.Column(db.Integer, db.ForeignKey('especialidad.id'), nullable=True)
    activo = db.Column(db.Boolean, default=True)
    ultimo_login = db.Column(db.DateTime, nullable=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    # Hardening seguridad (fase go-live)
    failed_attempts = db.Column(db.Integer, default=0, nullable=False, server_default='0')
    locked_until = db.Column(db.DateTime, nullable=True)
    must_change_password = db.Column(db.Boolean, default=False, nullable=False, server_default='false')
    alertas_stock = db.relationship('AlertaStock', backref='usuario', lazy=True)

class Curso(db.Model):
    """Curso/lectivo dentro de una especialidad. Ej: '3°A Electrónica'."""
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)  # ej: "3°A Electrónica" o "4°B"
    nivel = db.Column(db.String(20), nullable=True)     # ej: "3° Medio"
    letra = db.Column(db.String(5), nullable=True)      # ej: "A", "B"
    anio = db.Column(db.Integer, nullable=True)         # año lectivo
    especialidad_id = db.Column(db.Integer, db.ForeignKey('especialidad.id'), nullable=False)
    activo = db.Column(db.Boolean, default=True, nullable=False, server_default='1')
    # Cursos a cargo del pañol: máx. 2 con a_cargo=True por especialidad.
    # Identifican los grupos "propios" del pañol; el resto de alumnos son visitantes.
    a_cargo = db.Column(db.Boolean, default=False, nullable=False, server_default='0')
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    especialidad = db.relationship('Especialidad', backref='cursos', lazy=True)

    __table_args__ = (
        db.UniqueConstraint('nombre', 'especialidad_id', name='uq_curso_nombre_esp'),
    )

    @property
    def total_alumnos(self):
        return Estudiante.query.filter_by(curso_id=self.id, activo=True).count()


class Estudiante(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    rut_matricula = db.Column(db.String(50), unique=True, nullable=False)
    nombre = db.Column(db.String(100), nullable=False)
    curso = db.Column(db.String(50))  # legado: string libre (se mantiene por compatibilidad)
    curso_id = db.Column(db.Integer, db.ForeignKey('curso.id'), nullable=True)
    numero_lista = db.Column(db.Integer, nullable=True, index=True)  # N° lista dentro del curso
    codigo_barras = db.Column(db.String(50), unique=True, nullable=True, index=True)  # único global
    especialidad_id = db.Column(db.Integer, db.ForeignKey('especialidad.id'), nullable=False)
    email = db.Column(db.String(120), nullable=True)
    password_hash = db.Column(db.String(200), nullable=False)
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    especialidad = db.relationship('Especialidad', backref='estudiantes', lazy=True)
    curso_rel = db.relationship('Curso', backref='alumnos', lazy=True)

    # N° lista único dentro de un curso (puede repetirse entre cursos)
    __table_args__ = (
        db.UniqueConstraint('curso_id', 'numero_lista', name='uq_estud_curso_numlista'),
    )

    @property
    def tiene_deudas(self):
        return Prestamo.query.filter_by(estudiante_id=self.id, estado='Pendiente').count() > 0

    @property
    def curso_display(self):
        """Nombre legible del curso: usa el del FK si existe, si no el string legado."""
        if self.curso_rel:
            return self.curso_rel.nombre
        return self.curso or '—'

    @property
    def etiqueta_display(self):
        """Formato para dropdowns: 'N°lista — Nombre — Curso'."""
        partes = []
        if self.numero_lista is not None:
            partes.append(f"N°{self.numero_lista}")
        partes.append(self.nombre)
        if self.curso_display and self.curso_display != '—':
            partes.append(self.curso_display)
        return ' — '.join(partes)


def generar_codigo_barras_alumno():
    """Genera un código de barras único para un alumno: AL + timestamp + 3 dígitos."""
    import random
    base = "AL" + datetime.now().strftime('%y%m%d%H%M%S')
    # Reintenta si por azar colisiona (improbable)
    for _ in range(5):
        codigo = base + f"{random.randint(0, 999):03d}"
        if not Estudiante.query.filter_by(codigo_barras=codigo).first():
            return codigo
    # Fallback con uuid
    import uuid
    return "AL" + uuid.uuid4().hex[:14].upper()


class Item(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo_barras = db.Column(db.String(100), unique=True, nullable=True)
    nombre = db.Column(db.String(100), nullable=False)
    descripcion = db.Column(db.Text, nullable=True)
    especialidad_id = db.Column(db.Integer, db.ForeignKey('especialidad.id'), nullable=False)
    categoria = db.Column(db.String(50))
    cantidad_total = db.Column(db.Integer, default=0)
    cantidad_disponible = db.Column(db.Integer, default=0)
    cantidad_mermada = db.Column(db.Integer, default=0)
    cantidad_minima = db.Column(db.Integer, default=5)
    imagen_url = db.Column(db.String(500), default="")
    ubicacion = db.Column(db.String(200), default="Sin especificar")
    precio_unitario = db.Column(db.Float, default=0.0)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    fecha_ultima_modificacion = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    # Campos específicos Biblioteca (nullable, solo se llenan en items bibliográficos)
    autor = db.Column(db.String(200), nullable=True)
    isbn = db.Column(db.String(50), nullable=True, index=True)
    editorial = db.Column(db.String(200), nullable=True)
    anio_publicacion = db.Column(db.Integer, nullable=True)
    # Campos específicos INFORMATICA / DEPORTIVO
    marca = db.Column(db.String(100), nullable=True)
    modelo = db.Column(db.String(100), nullable=True)
    numero_serie = db.Column(db.String(100), nullable=True, index=True)
    estado = db.Column(db.String(50), nullable=True)  # Nuevo/Bueno/Regular/Reposición
    # Fecha de adquisición (todos los tipos, opcional)
    fecha_adquisicion = db.Column(db.Date, nullable=True)
    # Cálculo de desgaste por uso (INFORMATICA, PANOL_TP, DEPORTIVO)
    max_usos = db.Column(db.Integer, nullable=True)  # tope total de préstamos antes de reponer
    usos_actuales = db.Column(db.Integer, default=0, nullable=False,
                              server_default='0')
    # Desgaste/depreciación acumulada expresada en pesos chilenos (CLP)
    desgaste = db.Column(db.Float, default=0.0, nullable=False, server_default='0')

    @property
    def porcentaje_desgaste(self):
        """Devuelve % de desgaste basado en usos, o None si no aplica."""
        if self.max_usos and self.max_usos > 0:
            return min(100, int((self.usos_actuales or 0) * 100 / self.max_usos))
        return None

    @property
    def costo_total(self):
        """Costo total estimado del stock = costo unitario × cantidad total."""
        return (self.precio_unitario or 0.0) * (self.cantidad_total or 0)

class Prestamo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('item.id'), nullable=False)
    estudiante_id = db.Column(db.Integer, db.ForeignKey('estudiante.id'), nullable=False)
    profesor_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=True)
    # Nombre del profesor supervisor (modelo Profesor, gestionado por el instructor)
    profesor_nombre = db.Column(db.String(120), nullable=True)
    # Pañolero del día (estudiante designado que entregó el ítem). Opcional.
    panolero_dia_id = db.Column(db.Integer, db.ForeignKey('estudiante.id'), nullable=True)
    encargado = db.Column(db.String(100))
    cantidad = db.Column(db.Integer)
    cantidad_solicitada = db.Column(db.Integer, default=0)
    cantidad_mermada = db.Column(db.Integer, default=0)
    nombre_practica = db.Column(db.String(200), nullable=True)
    fecha_prestamo = db.Column(db.DateTime, default=datetime.utcnow)
    fecha_devolucion = db.Column(db.DateTime, nullable=True)
    estado = db.Column(db.String(20), default='Pendiente')
    item = db.relationship('Item', backref='prestamos_historial', lazy=True)
    # Como hay 2 FKs a estudiante (estudiante_id y panolero_dia_id), hay que
    # especificar explícitamente foreign_keys para evitar AmbiguousForeignKeysError.
    estudiante = db.relationship('Estudiante', backref='historial_solicitudes',
                                 foreign_keys=[estudiante_id], lazy=True)
    panolero_dia = db.relationship('Estudiante', backref='prestamos_atendidos',
                                   foreign_keys=[panolero_dia_id], lazy=True)
    profesor = db.relationship('Usuario', backref='prestamos_supervisados', lazy=True)
    # Campos específicos Biblioteca / préstamos con plazo
    fecha_devolucion_esperada = db.Column(db.DateTime, nullable=True)
    multa = db.Column(db.Float, default=0.0)

    @property
    def dias_atraso(self):
        if not self.fecha_devolucion_esperada or self.estado == 'Devuelto':
            return 0
        diff = (datetime.utcnow() - self.fecha_devolucion_esperada).days
        return max(0, diff)

class Auditoria(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=True)
    # nullable=True porque el admin no tiene especialidad asignada (puede operar sobre toda la BD)
    especialidad_id = db.Column(db.Integer, db.ForeignKey('especialidad.id'), nullable=True)
    accion = db.Column(db.String(100))
    tabla = db.Column(db.String(50))
    registro_id = db.Column(db.Integer)
    valores_anteriores = db.Column(db.Text)
    valores_nuevos = db.Column(db.Text)
    fecha = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    ip_address = db.Column(db.String(50))
    usuario = db.relationship('Usuario', backref='auditorias', lazy=True)
    especialidad = db.relationship('Especialidad', backref='auditorias', lazy=True)

class AlertaStock(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('item.id'), nullable=False)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    cantidad_minima = db.Column(db.Integer, default=5)
    cantidad_actual = db.Column(db.Integer)
    activa = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    fecha_ultima_alerta = db.Column(db.DateTime, nullable=True)
    item = db.relationship('Item', backref='alertas', lazy=True)

class OrdenTrabajo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    titulo = db.Column(db.String(200), nullable=False)
    descripcion = db.Column(db.Text, nullable=True)
    especialidad_id = db.Column(db.Integer, db.ForeignKey('especialidad.id'), nullable=False)
    profesional_cargo = db.Column(db.String(100), nullable=False)
    alumnos_cargo = db.Column(db.String(200), nullable=True)
    herramientas_utilizadas = db.Column(db.Text, nullable=True)
    repuestos_utilizados = db.Column(db.Text, nullable=True)
    profesor_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    estado = db.Column(db.String(50), default='Pendiente')
    especialidad = db.relationship('Especialidad', backref='ordenes_trabajo', lazy=True)
    profesor = db.relationship('Usuario', backref='ordenes_trabajo', lazy=True)

class ConfiguracionSistema(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    clave = db.Column(db.String(100), unique=True, nullable=False)
    valor = db.Column(db.Text)
    tipo = db.Column(db.String(20))
    descripcion = db.Column(db.Text)
    fecha_actualizacion = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class PrestamoExterno(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('item.id'), nullable=False)
    especialidad_id = db.Column(db.Integer, db.ForeignKey('especialidad.id'), nullable=False)
    cantidad = db.Column(db.Integer, nullable=False, default=1)
    es_alumno = db.Column(db.Boolean, default=False)
    persona_retira = db.Column(db.String(200), nullable=False)
    profesor_cargo = db.Column(db.String(200), nullable=True)
    especialidad_destino = db.Column(db.String(100), nullable=True)
    encargado = db.Column(db.String(100))
    fecha_prestamo = db.Column(db.DateTime, default=datetime.utcnow)
    fecha_devolucion = db.Column(db.DateTime, nullable=True)
    estado = db.Column(db.String(20), default='Activo')
    # 'prestamo' (se devuelve) o 'consumo' (descuenta del total, sin devolución; oficina)
    tipo_movimiento = db.Column(db.String(20), default='prestamo')
    item = db.relationship('Item', backref='prestamos_externos', lazy=True)
    especialidad = db.relationship('Especialidad', backref='prestamos_externos', lazy=True)


# ========================================================================
# PAÑOLEROS DEL DÍA — estudiantes designados que pueden entregar insumos
# ========================================================================
MAX_PANOLEROS_DIA = 6   # tope de pañoleros activos por especialidad
MAX_CURSOS_A_CARGO = 2  # tope de cursos a cargo por pañol (especialidad)


class PanoleroDesignado(db.Model):
    """Designación de un estudiante como pañolero del día para una especialidad.
    Persisten hasta que el encargado del pañol los reemplaza o limpia.
    Solo cuentan los registros con activo=True (máx. 6 por especialidad)."""
    __tablename__ = 'panolero_designado'
    id = db.Column(db.Integer, primary_key=True)
    estudiante_id = db.Column(db.Integer, db.ForeignKey('estudiante.id'),
                              nullable=False, index=True)
    especialidad_id = db.Column(db.Integer, db.ForeignKey('especialidad.id'),
                                nullable=False, index=True)
    designado_por_id = db.Column(db.Integer, db.ForeignKey('usuario.id'),
                                 nullable=True)
    fecha_designacion = db.Column(db.DateTime, default=datetime.utcnow,
                                  nullable=False)
    fecha_baja = db.Column(db.DateTime, nullable=True)
    activo = db.Column(db.Boolean, default=True, nullable=False,
                       server_default='true', index=True)
    estudiante = db.relationship('Estudiante', backref='designaciones_panolero', lazy=True)
    especialidad = db.relationship('Especialidad', backref='panoleros_designados', lazy=True)
    designado_por = db.relationship('Usuario', foreign_keys=[designado_por_id], lazy=True)


class Profesor(db.Model):
    """Profesor supervisor de un área. NO es una cuenta del sistema: es solo un
    nombre que el instructor registra para que aparezca como 'supervisor a cargo'
    en préstamos y órdenes de trabajo. Los gestiona el instructor de cada área."""
    __tablename__ = 'profesor'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(120), nullable=False)
    especialidad_id = db.Column(db.Integer, db.ForeignKey('especialidad.id'),
                                nullable=False, index=True)
    activo = db.Column(db.Boolean, default=True, nullable=False,
                       server_default='true', index=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    especialidad = db.relationship('Especialidad', backref='profesores_area', lazy=True)


def _panoleros_dia_activos(especialidad_id):
    """Devuelve las designaciones activas (objetos PanoleroDesignado COMPLETOS,
    no solo Estudiantes) para una especialidad. El template usa pd.estudiante.nombre
    y pd.id (para el form de quitar), así que necesita el PanoleroDesignado.
    Tolerante: si la tabla aún no existe, retorna []."""
    if not especialidad_id:
        return []
    try:
        designaciones = (PanoleroDesignado.query
                         .filter_by(especialidad_id=especialidad_id, activo=True)
                         .order_by(PanoleroDesignado.fecha_designacion.desc())
                         .limit(MAX_PANOLEROS_DIA)
                         .all())
        # Filtrar designaciones cuyo estudiante fue borrado (FK rota) para no romper la vista
        return [d for d in designaciones if d.estudiante is not None]
    except Exception as e:
        print(f"[WARN] _panoleros_dia_activos: {e}")
        return []


# 11. SYNC LOG (registro de cambios para sincronización entre nodos y admin central)
class SyncLog(db.Model):
    __tablename__ = 'sync_log'
    id = db.Column(db.Integer, primary_key=True)
    nodo_origen = db.Column(db.String(80), nullable=False, index=True)   # quién originó el cambio
    tabla = db.Column(db.String(50), nullable=False)                      # ej. 'item', 'prestamo'
    registro_id_local = db.Column(db.Integer, nullable=False)             # id en la BD local del nodo
    accion = db.Column(db.String(20), nullable=False)                     # 'crear', 'actualizar', 'eliminar'
    payload = db.Column(db.Text)                                          # JSON con el snapshot del registro
    fecha = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    push_status = db.Column(db.String(20), default='pendiente')           # 'pendiente', 'enviado', 'error'
    push_intentos = db.Column(db.Integer, default=0)
    push_error = db.Column(db.Text, nullable=True)
    sync_uuid = db.Column(db.String(64), unique=True, index=True)         # idempotencia

# ========== INICIALIZACIÓN DE BD ==========

def crear_especialidades_por_defecto():
    """17 áreas: 5 TP (campus único) + 6 áreas en 2 sedes cada una (norte/sur)."""
    especialidades = [
        # === 5 especialidades TP (sede única) ===
        {'nombre': 'Electrónica', 'color': '#FF6B6B'},
        {'nombre': 'Mecánica Automotriz', 'color': '#4ECDC4'},
        {'nombre': 'Mecánica Industrial', 'color': '#45B7D1'},
        {'nombre': 'Electricidad', 'color': '#FFA07A'},
        {'nombre': 'Gráfica', 'color': '#98D8C8'},
        # === 6 áreas duplicadas por sede ===
        {'nombre': 'ACLE Sede Norte', 'color': '#A78BFA',
         'descripcion': 'Actividades Curriculares de Libre Elección — Sede Norte'},
        {'nombre': 'ACLE Sede Sur', 'color': '#7C3AED',
         'descripcion': 'Actividades Curriculares de Libre Elección — Sede Sur'},
        {'nombre': 'Biblioteca Sede Norte', 'color': '#10B981',
         'descripcion': 'Libros y material bibliográfico — Sede Norte'},
        {'nombre': 'Biblioteca Sede Sur', 'color': '#047857',
         'descripcion': 'Libros y material bibliográfico — Sede Sur'},
        {'nombre': 'Informática Sede Norte', 'color': '#3B82F6',
         'descripcion': 'Sala/equipos de informática — Sede Norte'},
        {'nombre': 'Informática Sede Sur', 'color': '#1E40AF',
         'descripcion': 'Sala/equipos de informática — Sede Sur'},
        {'nombre': 'Educación Física Sede Norte', 'color': '#F97316',
         'descripcion': 'Implementos deportivos — Sede Norte'},
        {'nombre': 'Educación Física Sede Sur', 'color': '#C2410C',
         'descripcion': 'Implementos deportivos — Sede Sur'},
        {'nombre': 'Salas de Clase Sede Norte', 'color': '#EC4899',
         'descripcion': 'Recursos de salas de clase — Sede Norte'},
        {'nombre': 'Salas de Clase Sede Sur', 'color': '#BE185D',
         'descripcion': 'Recursos de salas de clase — Sede Sur'},
        {'nombre': 'Oficina Sede Norte', 'color': '#F59E0B',
         'descripcion': 'Implementos administrativos — Sede Norte'},
        {'nombre': 'Oficina Sede Sur', 'color': '#B45309',
         'descripcion': 'Implementos administrativos — Sede Sur'},
    ]
    for esp_data in especialidades:
        if not Especialidad.query.filter_by(nombre=esp_data['nombre']).first():
            db.session.add(Especialidad(
                nombre=esp_data['nombre'],
                color=esp_data['color'],
                descripcion=esp_data.get('descripcion', '')
            ))
    db.session.commit()

def crear_admin_central():
    if not Usuario.query.filter_by(username='admin_central').first():
        db.session.add(Usuario(
            nombre="Administrador Central", username="admin_central",
            password_hash=generate_password_hash("admin123"),
            rol="Admin", email="admin@colegio.local", especialidad_id=None,
            must_change_password=True,  # forzado a cambiar en primer login
        ))
        db.session.commit()

def crear_pañoleros_por_especialidad():
    import unicodedata
    for esp in Especialidad.query.all():
        nombre_norm = ''.join(c for c in unicodedata.normalize('NFD', esp.nombre)
                              if unicodedata.category(c) != 'Mn').lower().replace(' ', '_')
        username = f"pañolero_{nombre_norm}"
        email = f"pañolero.{nombre_norm}@colegio.local"
        existe = Usuario.query.filter(
            (Usuario.username == username) | (Usuario.email == email) |
            ((Usuario.rol == 'Pañolero') & (Usuario.especialidad_id == esp.id))
        ).first()
        if not existe:
            db.session.add(Usuario(
                nombre=f"Instructor {esp.nombre}", username=username,
                password_hash=generate_password_hash("pañol123"),
                rol="Pañolero", email=email, especialidad_id=esp.id,
                must_change_password=True,  # forzado a cambiar en primer login
            ))
    db.session.commit()

def _migrar_columnas_seguridad():
    """Añade columnas nuevas a usuario, especialidad e item si la BD ya existía.
    Idempotente — seguro de correr en cada arranque.
    """
    from sqlalchemy import inspect, text
    try:
        # 1) Asegurar que TODAS las tablas existan (curso, item, etc.). create_all es idempotente.
        db.create_all()

        insp = inspect(db.engine)
        tablas = set(insp.get_table_names())
        dialect = db.engine.dialect.name
        ts_type = 'TIMESTAMP' if dialect == 'postgresql' else 'DATETIME'
        date_type = 'DATE' if dialect == 'postgresql' else 'DATE'
        bool_default = 'FALSE' if dialect == 'postgresql' else '0'
        statements = []

        # === usuario: hardening de seguridad ===
        if 'usuario' in tablas:
            cols = {c['name'] for c in insp.get_columns('usuario')}
            if 'failed_attempts' not in cols:
                statements.append("ALTER TABLE usuario ADD COLUMN failed_attempts INTEGER NOT NULL DEFAULT 0")
            if 'locked_until' not in cols:
                statements.append(f"ALTER TABLE usuario ADD COLUMN locked_until {ts_type} NULL")
            if 'must_change_password' not in cols:
                statements.append(
                    f"ALTER TABLE usuario ADD COLUMN must_change_password BOOLEAN NOT NULL DEFAULT {bool_default}"
                )

        # === especialidad: tipo_area ===
        if 'especialidad' in tablas:
            cols = {c['name'] for c in insp.get_columns('especialidad')}
            if 'tipo_area' not in cols:
                statements.append(
                    "ALTER TABLE especialidad ADD COLUMN tipo_area VARCHAR(20) NOT NULL DEFAULT 'GENERAL'"
                )

        # === item: campos por tipo de área ===
        if 'item' in tablas:
            cols = {c['name'] for c in insp.get_columns('item')}
            if 'marca' not in cols:
                statements.append("ALTER TABLE item ADD COLUMN marca VARCHAR(100) NULL")
            if 'modelo' not in cols:
                statements.append("ALTER TABLE item ADD COLUMN modelo VARCHAR(100) NULL")
            if 'numero_serie' not in cols:
                statements.append("ALTER TABLE item ADD COLUMN numero_serie VARCHAR(100) NULL")
            if 'estado' not in cols:
                statements.append("ALTER TABLE item ADD COLUMN estado VARCHAR(50) NULL")
            if 'fecha_adquisicion' not in cols:
                statements.append(f"ALTER TABLE item ADD COLUMN fecha_adquisicion {date_type} NULL")
            if 'max_usos' not in cols:
                statements.append("ALTER TABLE item ADD COLUMN max_usos INTEGER NULL")
            if 'usos_actuales' not in cols:
                statements.append("ALTER TABLE item ADD COLUMN usos_actuales INTEGER NOT NULL DEFAULT 0")
            if 'desgaste' not in cols:
                statements.append("ALTER TABLE item ADD COLUMN desgaste FLOAT NOT NULL DEFAULT 0")

        # === prestamo: panolero_dia_id, profesor_nombre ===
        if 'prestamo' in tablas:
            cols = {c['name'] for c in insp.get_columns('prestamo')}
            if 'panolero_dia_id' not in cols:
                statements.append("ALTER TABLE prestamo ADD COLUMN panolero_dia_id INTEGER NULL")
            if 'profesor_nombre' not in cols:
                statements.append("ALTER TABLE prestamo ADD COLUMN profesor_nombre VARCHAR(120) NULL")

        # === estudiante: curso_id, email, numero_lista, codigo_barras ===
        if 'estudiante' in tablas:
            cols = {c['name'] for c in insp.get_columns('estudiante')}
            if 'curso_id' not in cols:
                statements.append("ALTER TABLE estudiante ADD COLUMN curso_id INTEGER NULL")
            if 'email' not in cols:
                statements.append("ALTER TABLE estudiante ADD COLUMN email VARCHAR(120) NULL")
            if 'numero_lista' not in cols:
                statements.append("ALTER TABLE estudiante ADD COLUMN numero_lista INTEGER NULL")
            if 'codigo_barras' not in cols:
                statements.append("ALTER TABLE estudiante ADD COLUMN codigo_barras VARCHAR(50) NULL")

        # === curso: a_cargo (flag de cursos a cargo del pañol, máx. 2 por especialidad) ===
        if 'curso' in tablas:
            cols = {c['name'] for c in insp.get_columns('curso')}
            if 'a_cargo' not in cols:
                statements.append(
                    f"ALTER TABLE curso ADD COLUMN a_cargo BOOLEAN NOT NULL DEFAULT {bool_default}"
                )

        if statements:
            with db.engine.begin() as conn:
                for s in statements:
                    try:
                        conn.execute(text(s))
                        print(f"[MIGRACION OK] {s}")
                    except Exception as exc:
                        # Loggear pero seguir con el resto. Asi no se aborta toda la migracion
                        # si una sola sentencia falla (ej: columna ya existe en otro orden).
                        print(f"[MIGRACION FALLO] {s}  ->  {exc}")
            print(f"[MIGRACION] Procesadas {len(statements)} sentencias")
    except Exception as e:
        # Esto SI es grave: imprimirlo con traceback para diagnosticar en logs
        import traceback
        print(f"[MIGRACION FATAL] {e}")
        traceback.print_exc()


# Mapeo de especialidad → tipo de área (se usa en seeders y en _asignar_tipo_area)
TIPO_AREA_POR_NOMBRE = {
    'Electrónica':                     'PANOL_TP',
    'Mecánica Automotriz':             'PANOL_TP',
    'Mecánica Industrial':             'PANOL_TP',
    'Electricidad':                    'PANOL_TP',
    'Gráfica':                         'PANOL_TP',
    'Biblioteca Sede Norte':           'BIBLIOTECA',
    'Biblioteca Sede Sur':             'BIBLIOTECA',
    'Educación Física Sede Norte':     'DEPORTIVO',
    'Educación Física Sede Sur':       'DEPORTIVO',
    'Informática Sede Norte':          'INFORMATICA',
    'Informática Sede Sur':            'INFORMATICA',
    'ACLE Sede Norte':                 'GENERAL',
    'ACLE Sede Sur':                   'GENERAL',
    'Salas de Clase Sede Norte':       'GENERAL',
    'Salas de Clase Sede Sur':         'GENERAL',
    'Oficina Sede Norte':              'GENERAL',
    'Oficina Sede Sur':                'GENERAL',
}


def _asignar_tipo_area():
    """Garantiza que cada especialidad tenga su tipo_area correcto. Idempotente."""
    cambios = 0
    for esp in Especialidad.query.all():
        tipo_esperado = TIPO_AREA_POR_NOMBRE.get(esp.nombre, 'GENERAL')
        if (esp.tipo_area or 'GENERAL') != tipo_esperado:
            esp.tipo_area = tipo_esperado
            cambios += 1
    if cambios:
        db.session.commit()
        print(f"[TIPO_AREA] {cambios} especialidades actualizadas")


def _flag_passwords_default():
    """Marca must_change_password=True en usuarios sembrados que aún tienen las
    contraseñas por defecto (admin123 / pañol123). Imprescindible para go-live cloud
    cuando la BD ya tenía esos usuarios creados antes de añadir el flag.
    """
    cambios = 0
    for u in Usuario.query.all():
        if u.must_change_password:
            continue
        # Admin con password default
        if u.username == 'admin_central' and check_password_hash(u.password_hash, 'admin123'):
            u.must_change_password = True
            cambios += 1
        # Pañoleros con password default
        elif u.rol == 'Pañolero' and check_password_hash(u.password_hash, 'pañol123'):
            u.must_change_password = True
            cambios += 1
    if cambios:
        db.session.commit()
        print(f"[SEC] {cambios} usuarios marcados para cambio obligatorio de contraseña")


with app.app_context():
    db.create_all()
    _migrar_columnas_seguridad()
    crear_especialidades_por_defecto()
    _asignar_tipo_area()
    crear_admin_central()
    crear_pañoleros_por_especialidad()
    _flag_passwords_default()
    print("✅ BD inicializada correctamente")
    print(f"✅ {Especialidad.query.count()} especialidades disponibles")
    print("✅ Admin central creado")
    print(f"✅ {Usuario.query.filter_by(rol='Pañolero').count()} pañoleros creados")


# ========================================================================
# SYNC: helpers para registrar cambios locales y consultarlos
# ========================================================================

import uuid as _uuid_mod

# Tablas que se sincronizan al admin central (la auditoría se conserva en cada nodo + admin)
TABLAS_SYNC = ('item', 'estudiante', 'prestamo', 'prestamo_externo',
               'orden_trabajo', 'usuario')


def _serializar_registro(obj):
    """Convierte un modelo SQLAlchemy a dict JSON-friendly. Solo columnas básicas."""
    if obj is None:
        return None
    out = {}
    for col in obj.__table__.columns:
        val = getattr(obj, col.name, None)
        if isinstance(val, datetime):
            out[col.name] = val.isoformat()
        else:
            out[col.name] = val
    return out


def registrar_cambio_sync(tabla, registro_id_local, accion, obj=None):
    """Registra un cambio local en SyncLog, listo para ser empujado al admin central.

    - Si este PC ES el admin central, no se registra (no necesita empujarse a sí mismo).
    - Si es un nodo cliente, se persiste un SyncLog con payload=snapshot del objeto.
    """
    if ES_ADMIN_CENTRAL:
        return  # el admin no se sincroniza consigo mismo
    if tabla not in TABLAS_SYNC:
        return
    payload = json.dumps(_serializar_registro(obj)) if obj else None
    entry = SyncLog(
        nodo_origen=NODO_ID,
        tabla=tabla,
        registro_id_local=registro_id_local,
        accion=accion,
        payload=payload,
        push_status='pendiente',
        sync_uuid=_uuid_mod.uuid4().hex,
    )
    db.session.add(entry)
    db.session.commit()


def _ack_sync_uuids(uuids_ok, uuids_error_map):
    """Marca SyncLogs como enviados/error después de un push. Solo en el nodo cliente."""
    if not uuids_ok and not uuids_error_map:
        return
    for uid in uuids_ok:
        log = SyncLog.query.filter_by(sync_uuid=uid).first()
        if log:
            log.push_status = 'enviado'
            log.push_intentos = (log.push_intentos or 0) + 1
            log.push_error = None
    for uid, err in uuids_error_map.items():
        log = SyncLog.query.filter_by(sync_uuid=uid).first()
        if log:
            log.push_status = 'error'
            log.push_intentos = (log.push_intentos or 0) + 1
            log.push_error = (err or '')[:500]
    db.session.commit()

# ========== DECORADORES ==========

def registrar_auditoria(accion, tabla, registro_id, valores_anteriores=None, valores_nuevos=None, especialidad_id=None):
    usuario_id = session.get('usuario_id')
    if especialidad_id is None and 'usuario_especialidad_id' in session:
        especialidad_id = session['usuario_especialidad_id']
    db.session.add(Auditoria(
        usuario_id=usuario_id, especialidad_id=especialidad_id,
        accion=accion, tabla=tabla, registro_id=registro_id,
        valores_anteriores=json.dumps(valores_anteriores) if valores_anteriores else None,
        valores_nuevos=json.dumps(valores_nuevos) if valores_nuevos else None,
        ip_address=request.remote_addr
    ))
    db.session.commit()

def login_requerido(f):
    @wraps(f)
    def w(*a, **kw):
        if 'usuario_id' not in session:
            flash("Debes iniciar sesión.")
            return redirect(url_for('login'))
        return f(*a, **kw)
    return w


@app.before_request
def _forzar_cambio_password_si_corresponde():
    """Si el usuario logueado tiene must_change_password=True, bloquear navegación
    a cualquier ruta excepto la de cambio de contraseña, logout, y assets estáticos."""
    if 'usuario_id' not in session:
        return
    if session.get('usuario_rol') == 'Estudiante':
        return  # estudiantes no tienen este flag
    rutas_permitidas = {'admin_cambiar_password', 'logout', 'static'}
    if request.endpoint in rutas_permitidas:
        return
    # Solo verificamos en BD si la sesión marca el flag — evita query por request.
    if not session.get('forzar_cambio_password'):
        # Refrescar desde BD una vez por sesión: si la BD dice que debe cambiar, redirigir.
        user = Usuario.query.get(session['usuario_id'])
        if user and user.must_change_password:
            session['forzar_cambio_password'] = True
        else:
            return
    flash("⚠️ Debes cambiar tu contraseña inicial antes de usar el sistema.")
    return redirect(url_for('admin_cambiar_password'))

def admin_requerido(f):
    @wraps(f)
    def w(*a, **kw):
        if session.get('usuario_rol') != 'Admin':
            flash("❌ Acceso denegado. Requiere Administrador.")
            return redirect(url_for('ver_inventario'))
        return f(*a, **kw)
    return w

def pañolero_o_admin(f):
    @wraps(f)
    def w(*a, **kw):
        if session.get('usuario_rol') not in ['Admin', 'Pañolero', 'Profesor']:
            flash("❌ Acceso denegado.")
            return redirect(url_for('ver_inventario'))
        return f(*a, **kw)
    return w

# ========== RUTAS BASICAS ==========

@app.route('/')
def index():
    if 'usuario_id' in session:
        if session.get('usuario_rol') == 'Admin':
            return redirect(url_for('dashboard_admin'))
        return redirect(url_for('ver_inventario'))
    return redirect(url_for('login'))

MAX_INTENTOS_LOGIN = 5
LOCKOUT_MINUTOS = 15


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        usuario = Usuario.query.filter_by(username=username).first()

        # 1) Bloqueo activo: rechazar incluso si la contraseña fuese correcta.
        if usuario and usuario.locked_until and usuario.locked_until > datetime.utcnow():
            restantes_min = int((usuario.locked_until - datetime.utcnow()).total_seconds() / 60) + 1
            flash(f"🔒 Cuenta bloqueada por intentos fallidos. Vuelve en {restantes_min} minutos.")
            return render_template('login.html')

        # 2) Login OK
        if usuario and usuario.activo and check_password_hash(usuario.password_hash, password):
            # Resetear contador de fallos al ingresar correctamente
            usuario.failed_attempts = 0
            usuario.locked_until = None

            session['usuario_id'] = usuario.id
            session['usuario_nombre'] = usuario.nombre
            session['usuario_rol'] = usuario.rol
            session['usuario_email'] = usuario.email
            if usuario.especialidad_id:
                session['usuario_especialidad_id'] = usuario.especialidad_id
                session['usuario_especialidad'] = usuario.especialidad_asignada.nombre
            usuario.ultimo_login = datetime.utcnow()
            db.session.commit()

            # Forzar cambio de contraseña en primer login
            if usuario.must_change_password:
                session['forzar_cambio_password'] = True
                flash("⚠️ Por seguridad, debes cambiar tu contraseña antes de continuar.")
                return redirect(url_for('admin_cambiar_password'))

            flash(f"✅ Bienvenido {usuario.nombre}")
            return redirect(url_for('index'))

        # 3) Login de estudiante (por RUT) — sin lockout (lo añadiremos si hace falta)
        estudiante = Estudiante.query.filter_by(rut_matricula=username).first()
        if estudiante and estudiante.activo and check_password_hash(estudiante.password_hash, password):
            session['usuario_id'] = estudiante.id
            session['usuario_nombre'] = estudiante.nombre
            session['usuario_rol'] = 'Estudiante'
            session['usuario_especialidad_id'] = estudiante.especialidad_id
            session['usuario_especialidad'] = estudiante.especialidad.nombre
            flash(f"✅ Bienvenido {estudiante.nombre}")
            return redirect(url_for('ver_inventario'))

        # 4) Credenciales incorrectas: si el usuario existe, incrementar contador.
        if usuario:
            usuario.failed_attempts = (usuario.failed_attempts or 0) + 1
            if usuario.failed_attempts >= MAX_INTENTOS_LOGIN:
                usuario.locked_until = datetime.utcnow() + timedelta(minutes=LOCKOUT_MINUTOS)
                usuario.failed_attempts = 0
                db.session.commit()
                flash(f"🔒 Demasiados intentos fallidos. Cuenta bloqueada por {LOCKOUT_MINUTOS} minutos.")
            else:
                restantes = MAX_INTENTOS_LOGIN - usuario.failed_attempts
                db.session.commit()
                flash(f"❌ Credenciales incorrectas. Te quedan {restantes} intento(s) antes del bloqueo.")
        else:
            # No revelar si el usuario existe o no: mensaje genérico.
            flash("❌ Credenciales incorrectas.")
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash("✅ Sesión cerrada.")
    return redirect(url_for('login'))

@app.route('/dashboard_admin')
@login_requerido
@admin_requerido
def dashboard_admin():
    # Al volver al panel, limpiamos cualquier área temporal que quedó del último ingreso
    session.pop('usuario_especialidad_id', None)
    session.pop('usuario_especialidad', None)
    session.pop('admin_viendo_especialidad', None)

    especialidades = Especialidad.query.filter_by(activa=True).all()
    total_items = Item.query.count()
    total_stock = sum(i.cantidad_total for i in Item.query.all())
    total_estudiantes = Estudiante.query.count()
    prestamos_activos = Prestamo.query.filter_by(estado='Pendiente').count()
    stats_por_especialidad = []
    for esp in especialidades:
        items_esp = Item.query.filter_by(especialidad_id=esp.id).all()
        prest_esp = Prestamo.query.join(Item).filter(
            Item.especialidad_id == esp.id, Prestamo.estado == 'Pendiente'
        ).count()
        stats_por_especialidad.append({
            'id': esp.id,
            'especialidad': esp.nombre,
            'color': esp.color or '#2563eb',
            'items_totales': len(items_esp),
            'stock_total': sum(i.cantidad_total for i in items_esp),
            'estudiantes': Estudiante.query.filter_by(especialidad_id=esp.id, activo=True).count(),
            'prestamos_activos': prest_esp,
        })
    return render_template('dashboard_admin.html',
                           especialidades=especialidades,
                           total_items=total_items, total_stock=total_stock,
                           total_estudiantes=total_estudiantes,
                           prestamos_activos=prestamos_activos,
                           stats_por_especialidad=stats_por_especialidad)

def _calcular_practicas_resumen(especialidad_id):
    practicas = {}
    qs = Prestamo.query.join(Item).filter(
        Item.especialidad_id == especialidad_id,
        Prestamo.nombre_practica.isnot(None),
        Prestamo.nombre_practica != ''
    ).all()
    for p in qs:
        n = p.nombre_practica
        if n not in practicas:
            practicas[n] = {'nombre_practica': n, 'fecha': p.fecha_prestamo,
                            'herr_detalles': [], 'mat_detalles': [],
                            'comp_detalles': [], 'fung_detalles': []}
        cat = (p.item.categoria or '').lower()
        linea = f"{p.item.nombre} x{p.cantidad}"
        if 'herr' in cat: practicas[n]['herr_detalles'].append(linea)
        elif 'mat' in cat: practicas[n]['mat_detalles'].append(linea)
        elif 'comp' in cat: practicas[n]['comp_detalles'].append(linea)
        else: practicas[n]['fung_detalles'].append(linea)
    return sorted(practicas.values(), key=lambda x: x['fecha'], reverse=True)

@app.route('/inventario')
@login_requerido
def ver_inventario():
    rol = session.get('usuario_rol')
    if rol == 'Admin':
        # Admin puede pasar ?especialidad_id=X para ver inventario de un área
        esp_query = request.args.get('especialidad_id', type=int)
        if not esp_query:
            return redirect(url_for('dashboard_admin'))
        esp_obj = Especialidad.query.get(esp_query)
        if not esp_obj:
            flash("❌ Especialidad no encontrada.")
            return redirect(url_for('dashboard_admin'))
        especialidad_id = esp_query
        # Setear sesión temporal para que los POSTs (agregar, prestar, etc.) sepan el área
        session['admin_viendo_especialidad'] = esp_obj.nombre
        session['usuario_especialidad_id'] = esp_query
        session['usuario_especialidad'] = esp_obj.nombre
    else:
        especialidad_id = session.get('usuario_especialidad_id')
    items = Item.query.filter_by(especialidad_id=especialidad_id) \
        .order_by(Item.categoria.asc(), Item.nombre.asc()).all()
    estudiantes = Estudiante.query.filter_by(especialidad_id=especialidad_id, activo=True).all()
    prestamos = Prestamo.query.join(Item).filter(
        Item.especialidad_id == especialidad_id
    ).order_by(Prestamo.fecha_prestamo.desc()).limit(50).all()
    prestamos_externos = PrestamoExterno.query.filter_by(
        especialidad_id=especialidad_id
    ).order_by(PrestamoExterno.fecha_prestamo.desc()).limit(50).all()
    ordenes_trabajo = OrdenTrabajo.query.filter_by(
        especialidad_id=especialidad_id
    ).order_by(OrdenTrabajo.fecha_creacion.desc()).limit(50).all()
    profesores = Usuario.query.filter(
        Usuario.rol.in_(['Profesor', 'Pañolero', 'Admin']),
        Usuario.activo == True
    ).all()
    usuarios_sistema = Usuario.query.filter(
        (Usuario.especialidad_id == especialidad_id) | (Usuario.rol == 'Admin')
    ).all()
    practicas_resumen = _calcular_practicas_resumen(especialidad_id)
    alertas = AlertaStock.query.filter_by(usuario_id=session.get('usuario_id'), activa=True).all()
    total_stock = sum(i.cantidad_total for i in items)
    prestamos_activos = Prestamo.query.join(Item).filter(
        Item.especialidad_id == especialidad_id, Prestamo.estado == 'Pendiente'
    ).count()
    # Tipo de área: define qué campos del formulario y columnas de la lista se muestran
    esp_obj_actual = Especialidad.query.get(especialidad_id)
    tipo_area = (esp_obj_actual.tipo_area or 'GENERAL') if esp_obj_actual else 'GENERAL'
    # Pañoleros del día activos para esta especialidad (máx. 6)
    panoleros_dia = _panoleros_dia_activos(especialidad_id) if especialidad_id else []
    # Cursos disponibles en esta especialidad (para datalist/dropdowns)
    try:
        cursos_disponibles = Curso.query.filter_by(
            especialidad_id=especialidad_id, activo=True
        ).order_by(Curso.nombre.asc()).all()
    except Exception:
        cursos_disponibles = []
    # Cursos a cargo del pañol (máx. 2) y alumnos visitantes de otras especialidades
    cursos_a_cargo = _cursos_a_cargo(especialidad_id) if especialidad_id else []
    alumnos_visitantes = _alumnos_visitantes(especialidad_id) if especialidad_id else []
    profesores_area = _profesores_activos(especialidad_id) if especialidad_id else []
    return render_template('inventario.html',
                           items=items, estudiantes=estudiantes, prestamos=prestamos,
                           prestamos_externos=prestamos_externos,
                           ordenes_trabajo=ordenes_trabajo, profesores=profesores,
                           profesores_area=profesores_area,
                           usuarios_sistema=usuarios_sistema,
                           practicas_resumen=practicas_resumen,
                           total_stock=total_stock,
                           prestamos_activos=prestamos_activos,
                           alertas=alertas,
                           tipo_area=tipo_area,
                           panoleros_dia=panoleros_dia,
                           cursos_disponibles=cursos_disponibles,
                           cursos_a_cargo=cursos_a_cargo,
                           alumnos_visitantes=alumnos_visitantes,
                           max_panoleros_dia=MAX_PANOLEROS_DIA,
                           max_cursos_a_cargo=MAX_CURSOS_A_CARGO,
                           especialidad=(session.get('admin_viendo_especialidad')
                                          if session.get('usuario_rol') == 'Admin'
                                          else session.get('usuario_especialidad')),
                           especialidades_disponibles=Especialidad.query.order_by(Especialidad.nombre).all())

@app.route('/auditoria')
@login_requerido
def ver_auditoria():
    rol = session.get('usuario_rol')
    # Filtros opcionales por query string
    f_esp = request.args.get('especialidad_id', type=int)
    f_user = (request.args.get('usuario') or '').strip()
    f_accion = (request.args.get('accion') or '').strip()
    f_desde = request.args.get('desde')
    f_hasta = request.args.get('hasta')

    q = Auditoria.query
    if rol != 'Admin':
        q = q.filter_by(especialidad_id=session.get('usuario_especialidad_id'))
    if f_esp:
        q = q.filter_by(especialidad_id=f_esp)
    if f_accion:
        q = q.filter(Auditoria.accion.ilike(f"%{f_accion}%"))
    if f_user:
        q = q.join(Usuario).filter(
            db.or_(Usuario.username.ilike(f"%{f_user}%"),
                   Usuario.nombre.ilike(f"%{f_user}%"))
        )
    if f_desde:
        try:
            q = q.filter(Auditoria.fecha >= datetime.fromisoformat(f_desde))
        except Exception: pass
    if f_hasta:
        try:
            q = q.filter(Auditoria.fecha <= datetime.fromisoformat(f_hasta))
        except Exception: pass

    logs = q.order_by(Auditoria.fecha.desc()).limit(500).all()
    especialidades = Especialidad.query.filter_by(activa=True).all()
    return render_template('auditoria.html', logs=logs,
                           especialidades=especialidades,
                           f_esp=f_esp, f_user=f_user, f_accion=f_accion,
                           f_desde=f_desde, f_hasta=f_hasta,
                           es_admin=(rol == 'Admin'))

# ========== ITEMS ==========

@app.route('/agregar', methods=['POST'])
@login_requerido
@pañolero_o_admin
def agregar_item():
    especialidad_id = session.get('usuario_especialidad_id')
    codigo = request.form.get('codigo_barras', '').strip() or datetime.now().strftime('%y%m%d%H%M%S') + "0"
    cantidad = int(request.form.get('cantidad', 1))
    imagen = request.form.get('imagen_url', '').strip()
    ubicacion = request.form.get('ubicacion', 'Sin especificar').strip()
    nombre = request.form.get('nombre', '').strip()
    categoria = request.form.get('categoria', '').strip()

    # Campos opcionales tipo-específicos
    autor = request.form.get('autor', '').strip() or None
    isbn = request.form.get('isbn', '').strip() or None
    editorial = request.form.get('editorial', '').strip() or None
    anio_pub = request.form.get('anio_publicacion', type=int)
    marca = request.form.get('marca', '').strip() or None
    modelo = request.form.get('modelo', '').strip() or None
    numero_serie = request.form.get('numero_serie', '').strip() or None
    estado = request.form.get('estado', '').strip() or None
    max_usos = request.form.get('max_usos', type=int)
    fecha_adq_raw = request.form.get('fecha_adquisicion', '').strip()
    fecha_adq = None
    if fecha_adq_raw:
        try:
            fecha_adq = datetime.fromisoformat(fecha_adq_raw).date()
        except Exception:
            fecha_adq = None
    # Costos y desgaste (en pesos chilenos, opcionales)
    descripcion = request.form.get('descripcion', '').strip() or None
    precio_unitario = request.form.get('precio_unitario', type=float) or 0.0
    desgaste = request.form.get('desgaste', type=float) or 0.0

    item_existente = Item.query.filter_by(codigo_barras=codigo, especialidad_id=especialidad_id).first()
    nuevo_item = None
    if item_existente:
        item_existente.cantidad_total += cantidad
        item_existente.cantidad_disponible += cantidad
        if imagen: item_existente.imagen_url = imagen
        if ubicacion: item_existente.ubicacion = ubicacion
        # Solo actualizar campos opcionales si vienen con valor (no sobreescribir con vacío)
        if autor: item_existente.autor = autor
        if isbn: item_existente.isbn = isbn
        if editorial: item_existente.editorial = editorial
        if anio_pub: item_existente.anio_publicacion = anio_pub
        if marca: item_existente.marca = marca
        if modelo: item_existente.modelo = modelo
        if numero_serie: item_existente.numero_serie = numero_serie
        if estado: item_existente.estado = estado
        if max_usos: item_existente.max_usos = max_usos
        if fecha_adq: item_existente.fecha_adquisicion = fecha_adq
        if descripcion: item_existente.descripcion = descripcion
        if precio_unitario: item_existente.precio_unitario = precio_unitario
        if desgaste: item_existente.desgaste = desgaste
    else:
        nuevo_item = Item(codigo_barras=codigo, nombre=nombre, categoria=categoria,
                          descripcion=descripcion,
                          especialidad_id=especialidad_id,
                          cantidad_total=cantidad, cantidad_disponible=cantidad,
                          imagen_url=imagen, ubicacion=ubicacion,
                          precio_unitario=precio_unitario, desgaste=desgaste,
                          autor=autor, isbn=isbn, editorial=editorial,
                          anio_publicacion=anio_pub,
                          marca=marca, modelo=modelo, numero_serie=numero_serie,
                          estado=estado, fecha_adquisicion=fecha_adq,
                          max_usos=max_usos)
        db.session.add(nuevo_item)
    db.session.commit()
    rid = item_existente.id if item_existente else nuevo_item.id
    registrar_auditoria('crear', 'Item', rid, valores_nuevos={'nombre': nombre, 'cantidad': cantidad})
    registrar_cambio_sync('item', rid, 'crear' if not item_existente else 'actualizar', item_existente or nuevo_item)
    flash(f"✅ Ítem '{nombre}' agregado.")
    return redirect(url_for('ver_inventario'))

@app.route('/eliminar/<int:item_id>', methods=['POST'])
@login_requerido
@pañolero_o_admin
def eliminar_item(item_id):
    item = Item.query.get_or_404(item_id)
    if session.get('usuario_rol') != 'Admin' and item.especialidad_id != session.get('usuario_especialidad_id'):
        flash("❌ Sin permiso.")
        return redirect(url_for('ver_inventario'))
    if Prestamo.query.filter_by(item_id=item.id, estado='Pendiente').first() or \
       PrestamoExterno.query.filter_by(item_id=item.id, estado='Activo').first():
        flash(f"⚠️ '{item.nombre}' tiene préstamos pendientes.")
        return redirect(url_for('ver_inventario'))
    registrar_auditoria('eliminar', 'Item', item.id,
                        valores_anteriores={'nombre': item.nombre, 'codigo': item.codigo_barras})
    registrar_cambio_sync('item', item.id, 'eliminar', item)
    db.session.delete(item)
    db.session.commit()
    flash(f"✅ Ítem '{item.nombre}' eliminado.")
    return redirect(url_for('ver_inventario'))

@app.route('/editar_item/<int:item_id>', methods=['POST'])
@login_requerido
@pañolero_o_admin
def editar_item(item_id):
    item = Item.query.get_or_404(item_id)
    if session.get('usuario_rol') != 'Admin' and item.especialidad_id != session.get('usuario_especialidad_id'):
        flash("❌ Sin permiso.")
        return redirect(url_for('ver_inventario'))
    item.nombre = request.form.get('nombre', item.nombre).strip()
    item.categoria = request.form.get('categoria', item.categoria)
    item.ubicacion = request.form.get('ubicacion', item.ubicacion)
    item.cantidad_minima = int(request.form.get('cantidad_minima') or item.cantidad_minima)

    # Imagen y descripción (solo se cambian si vienen valores)
    nueva_img = request.form.get('imagen_url', '').strip()
    if nueva_img:
        item.imagen_url = nueva_img
    nueva_desc = request.form.get('descripcion', None)
    if nueva_desc is not None and nueva_desc.strip() != '':
        item.descripcion = nueva_desc.strip()

    # Costo unitario y desgaste en pesos (acepta vacío = no cambia)
    precio_raw = request.form.get('precio_unitario', '').strip()
    if precio_raw != '':
        try:
            item.precio_unitario = float(precio_raw)
        except ValueError:
            pass
    desg_raw = request.form.get('desgaste', '').strip()
    if desg_raw != '':
        try:
            item.desgaste = float(desg_raw)
        except ValueError:
            pass

    # Fecha de adquisición (opcional)
    fecha_adq_raw = request.form.get('fecha_adquisicion', '').strip()
    if fecha_adq_raw:
        try:
            item.fecha_adquisicion = datetime.fromisoformat(fecha_adq_raw).date()
        except Exception:
            pass

    # Ajuste de stock: soporta dos formas
    # 1) cantidad_total → setear total absoluto
    # 2) ajuste_cantidad → delta a sumar/restar
    nueva = request.form.get('cantidad_total')
    ajuste = request.form.get('ajuste_cantidad')
    if nueva is not None and nueva != '':
        diff = int(nueva) - item.cantidad_total
        item.cantidad_total = int(nueva)
        item.cantidad_disponible = max(0, item.cantidad_disponible + diff)
    elif ajuste is not None and ajuste != '' and ajuste != '0':
        try:
            d = int(ajuste)
            item.cantidad_total = max(0, item.cantidad_total + d)
            item.cantidad_disponible = max(0, item.cantidad_disponible + d)
        except ValueError:
            pass
    db.session.commit()
    registrar_auditoria('actualizar', 'Item', item.id, valores_nuevos={'nombre': item.nombre})
    registrar_cambio_sync('item', item.id, 'actualizar', item)
    flash(f"✅ Ítem actualizado.")
    return redirect(url_for('ver_inventario'))

# ========== ESTUDIANTES ==========

@app.route('/agregar_estudiante', methods=['POST'])
@login_requerido
@pañolero_o_admin
def agregar_estudiante():
    rut = request.form.get('rut_matricula', '').strip()
    nombre = request.form.get('nombre', '').strip()
    curso = request.form.get('curso', '').strip()
    especialidad_id = session.get('usuario_especialidad_id') or request.form.get('especialidad_id', type=int)
    if not rut or not nombre:
        flash("❌ RUT y nombre obligatorios.")
        return redirect(url_for('ver_inventario'))
    if Estudiante.query.filter_by(rut_matricula=rut).first():
        flash(f"⚠️ RUT {rut} ya existe.")
        return redirect(url_for('ver_inventario'))
    nuevo = Estudiante(rut_matricula=rut, nombre=nombre, curso=curso,
                       especialidad_id=especialidad_id,
                       password_hash=generate_password_hash(rut), activo=True)
    db.session.add(nuevo); db.session.commit()
    registrar_auditoria('crear', 'Estudiante', nuevo.id,
                        valores_nuevos={'rut': rut, 'nombre': nombre, 'curso': curso})
    registrar_cambio_sync('estudiante', nuevo.id, 'crear', nuevo)
    flash(f"✅ Estudiante {nombre} agregado.")
    return redirect(url_for('ver_inventario'))

@app.route('/cargar_alumnos_excel', methods=['POST'])
@login_requerido
@pañolero_o_admin
def cargar_alumnos_excel():
    """Carga masiva de alumnos desde Excel.

    Formato simplificado (2 columnas en orden fijo):
      Col 1 → N° de lista  (entero)
      Col 2 → Nombre completo (obligatorio)

    El CURSO al que pertenecen los alumnos se elige en un input del formulario
    (campo 'curso_destino'). Todos los alumnos de la planilla quedan asignados
    a ese curso. Si el curso no existe, se crea.

    Reglas:
      - Solo el pañolero crea alumnos en SU especialidad.
      - Identidad del alumno = curso + N° lista (puede repetirse N° entre cursos).
      - Si ya existe un alumno con ese (curso, N° lista), se actualiza su nombre.
      - Cada alumno obtiene un código de barras autogenerado (único global).
      - Contraseña inicial = código de barras del alumno.
    """
    archivo = request.files.get('archivo_excel')
    if not archivo or archivo.filename == '':
        flash("❌ No subiste ningún archivo.")
        return redirect(url_for('ver_inventario'))

    especialidad_id = session.get('usuario_especialidad_id')
    if not especialidad_id:
        flash("❌ Tu cuenta no tiene una especialidad asignada.")
        return redirect(url_for('ver_inventario'))

    nombre_curso = (request.form.get('curso_destino') or '').strip()
    if not nombre_curso:
        flash("❌ Debes indicar el curso al que pertenecen estos alumnos.")
        return redirect(url_for('ver_inventario'))

    try:
        df = pd.read_excel(archivo, header=None, dtype=object, sheet_name=0)
    except Exception as e:
        flash(f"❌ No pude leer el Excel: {e}")
        return redirect(url_for('ver_inventario'))

    if len(df) == 0:
        flash("⚠️ El Excel está vacío.")
        return redirect(url_for('ver_inventario'))

    # Detectar cabecera: si la primera fila col 1 (nombre) no es nombre real ni la col 0 (N°) es entero, saltar
    inicio = 0
    primera = df.iloc[0]
    try:
        # Si la primera celda no es número, probablemente es cabecera ("N° lista" / "Nombre")
        int(float(str(primera[0]).strip()))
    except (ValueError, TypeError, IndexError):
        inicio = 1

    def col(fila, i, default=''):
        try:
            v = fila[i]
        except (IndexError, KeyError):
            return default
        if v is None:
            return default
        if isinstance(v, float) and pd.isna(v):
            return default
        s = str(v).strip()
        if not s or s.lower() == 'nan':
            return default
        return s

    # Get-or-create del curso destino UNA sola vez
    curso_obj = Curso.query.filter_by(nombre=nombre_curso, especialidad_id=especialidad_id).first()
    curso_nuevo = False
    if not curso_obj:
        import re
        m = re.match(r'^\s*(\d+)\s*[°º]?\s*([A-Za-z])?', nombre_curso)
        nivel = f"{m.group(1)}° Medio" if (m and m.group(1)) else None
        letra = m.group(2).upper() if (m and m.group(2)) else None
        curso_obj = Curso(nombre=nombre_curso, nivel=nivel, letra=letra,
                          anio=datetime.now().year,
                          especialidad_id=especialidad_id, activo=True)
        db.session.add(curso_obj)
        db.session.flush()
        curso_nuevo = True

    # Auto-marcar como a_cargo si todavía hay slot libre (máx 2)
    if not curso_obj.a_cargo:
        try:
            ya_a_cargo = Curso.query.filter_by(
                especialidad_id=especialidad_id, activo=True, a_cargo=True
            ).count()
            if ya_a_cargo < MAX_CURSOS_A_CARGO:
                curso_obj.a_cargo = True
                db.session.flush()
        except Exception as e:
            print(f"[WARN] auto-marcar a_cargo: {e}")

    creados = actualizados = 0
    errores = []

    for idx in range(inicio, len(df)):
        fila = df.iloc[idx]
        n_lista_str = col(fila, 0)
        nombre = col(fila, 1)

        if not nombre:
            if n_lista_str:
                errores.append(f"Fila {idx + 1}: falta nombre")
            continue

        # N° lista
        try:
            n_lista = int(float(n_lista_str)) if n_lista_str else None
        except Exception:
            errores.append(f"Fila {idx + 1}: N° lista inválido ({n_lista_str})")
            continue

        # Buscar existente por (curso_id, numero_lista). Si no, crear nuevo.
        existente = None
        if n_lista is not None:
            existente = Estudiante.query.filter_by(curso_id=curso_obj.id,
                                                   numero_lista=n_lista).first()

        if existente:
            existente.nombre = nombre
            existente.curso = nombre_curso
            existente.curso_id = curso_obj.id
            # Si no tiene código de barras, generarle uno
            if not existente.codigo_barras:
                existente.codigo_barras = generar_codigo_barras_alumno()
            registrar_cambio_sync('estudiante', existente.id, 'actualizar', existente)
            actualizados += 1
        else:
            codigo_alumno = generar_codigo_barras_alumno()
            # rut_matricula se mantiene como identificador único interno; usamos el código de barras
            nuevo = Estudiante(
                rut_matricula=codigo_alumno,  # idéntico al codigo_barras para mantener unicidad
                codigo_barras=codigo_alumno,
                nombre=nombre,
                numero_lista=n_lista,
                curso=nombre_curso,
                curso_id=curso_obj.id,
                especialidad_id=especialidad_id,
                password_hash=generate_password_hash(codigo_alumno),
                activo=True,
            )
            db.session.add(nuevo)
            db.session.flush()
            registrar_cambio_sync('estudiante', nuevo.id, 'crear', nuevo)
            creados += 1

    db.session.commit()
    registrar_auditoria('importar', 'Estudiante', 0,
                        valores_nuevos={'creados': creados,
                                        'actualizados': actualizados,
                                        'curso': nombre_curso,
                                        'curso_nuevo': curso_nuevo,
                                        'errores': len(errores)})

    extra = " (curso recién creado)" if curso_nuevo else ""
    msg = (f"✅ Carga de alumnos a {nombre_curso}{extra}: "
           f"{creados} nuevo(s), {actualizados} actualizado(s).")
    if errores:
        msg += f" ⚠️ {len(errores)} fila(s) con error: " + " | ".join(errores[:3])
    flash(msg)
    return redirect(url_for('ver_inventario'))


@app.route('/editar_alumno/<int:est_id>', methods=['POST'])
@login_requerido
@pañolero_o_admin
def editar_alumno(est_id):
    """Permite editar nombre, N° lista y curso de un alumno.
    Solo el pañolero de su misma especialidad puede editarlo (o el Admin)."""
    est = Estudiante.query.get_or_404(est_id)
    if session.get('usuario_rol') != 'Admin' and est.especialidad_id != session.get('usuario_especialidad_id'):
        flash("❌ Sin permiso para editar a este alumno.")
        return redirect(url_for('ver_inventario'))

    nombre = (request.form.get('nombre') or '').strip()
    if nombre:
        est.nombre = nombre

    n_lista_raw = (request.form.get('numero_lista') or '').strip()
    if n_lista_raw:
        try:
            est.numero_lista = int(n_lista_raw)
        except ValueError:
            pass

    nombre_curso = (request.form.get('curso_destino') or '').strip()
    if nombre_curso:
        c = Curso.query.filter_by(nombre=nombre_curso,
                                  especialidad_id=est.especialidad_id).first()
        if not c:
            c = Curso(nombre=nombre_curso, especialidad_id=est.especialidad_id,
                      anio=datetime.now().year, activo=True)
            db.session.add(c)
            db.session.flush()
        est.curso_id = c.id
        est.curso = nombre_curso

    activo_val = request.form.get('activo')
    if activo_val is not None:
        est.activo = activo_val in ('1', 'true', 'on', 'si', 'sí')

    db.session.commit()
    registrar_auditoria('actualizar', 'Estudiante', est.id,
                        valores_nuevos={'nombre': est.nombre,
                                        'numero_lista': est.numero_lista,
                                        'curso': est.curso})
    registrar_cambio_sync('estudiante', est.id, 'actualizar', est)
    flash(f"✅ Alumno {est.nombre} actualizado.")
    return redirect(url_for('ver_inventario'))


@app.route('/regenerar_codigo_alumno/<int:est_id>', methods=['POST'])
@login_requerido
@pañolero_o_admin
def regenerar_codigo_alumno(est_id):
    """Genera un nuevo código de barras para el alumno (útil si pierde el carnet)."""
    est = Estudiante.query.get_or_404(est_id)
    if session.get('usuario_rol') != 'Admin' and est.especialidad_id != session.get('usuario_especialidad_id'):
        flash("❌ Sin permiso.")
        return redirect(url_for('ver_inventario'))
    est.codigo_barras = generar_codigo_barras_alumno()
    db.session.commit()
    registrar_auditoria('actualizar', 'Estudiante', est.id,
                        valores_nuevos={'accion': 'regenerar_codigo',
                                        'nuevo_codigo': est.codigo_barras})
    registrar_cambio_sync('estudiante', est.id, 'actualizar', est)
    flash(f"✅ Nuevo código de {est.nombre}: {est.codigo_barras}")
    return redirect(url_for('ver_inventario'))


@app.route('/etiqueta_alumno/<int:est_id>')
@login_requerido
def etiqueta_alumno(est_id):
    """Página imprimible con el carnet (código de barras) del alumno."""
    est = Estudiante.query.get_or_404(est_id)
    if session.get('usuario_rol') == 'Pañolero' and est.especialidad_id != session.get('usuario_especialidad_id'):
        flash("❌ Sin permiso.")
        return redirect(url_for('ver_inventario'))
    # Si no tiene código aún, generarlo
    if not est.codigo_barras:
        est.codigo_barras = generar_codigo_barras_alumno()
        db.session.commit()
    return render_template('etiqueta_alumno.html', alumno=est)


@app.route('/api/item/<codigo>')
@login_requerido
def api_buscar_item(codigo):
    """Devuelve JSON con datos del ítem si el código existe en la especialidad
    del usuario. Lo usa el formulario de agregar para autocompletar al escanear."""
    codigo = (codigo or '').strip()
    if not codigo:
        return jsonify({'found': False})
    especialidad_id = session.get('usuario_especialidad_id')
    q = Item.query.filter_by(codigo_barras=codigo)
    if especialidad_id and session.get('usuario_rol') != 'Admin':
        q = q.filter_by(especialidad_id=especialidad_id)
    item = q.first()
    if not item:
        return jsonify({'found': False})
    return jsonify({
        'found': True,
        'id': item.id,
        'nombre': item.nombre,
        'descripcion': item.descripcion or '',
        'categoria': item.categoria or '',
        'ubicacion': item.ubicacion or '',
        'imagen_url': item.imagen_url or '',
        'precio_unitario': item.precio_unitario or 0,
        'desgaste': item.desgaste or 0,
        'marca': item.marca or '',
        'modelo': item.modelo or '',
        'numero_serie': item.numero_serie or '',
        'estado': item.estado or '',
        'cantidad_total': item.cantidad_total or 0,
        'cantidad_disponible': item.cantidad_disponible or 0,
    })


@app.route('/eliminar_estudiante/<int:est_id>', methods=['POST'])
@login_requerido
@pañolero_o_admin
def eliminar_estudiante(est_id):
    est = Estudiante.query.get_or_404(est_id)
    if Prestamo.query.filter_by(estudiante_id=est.id, estado='Pendiente').first():
        flash(f"⚠️ {est.nombre} tiene préstamos pendientes.")
        return redirect(url_for('ver_inventario'))
    est.activo = False
    db.session.commit()
    registrar_auditoria('eliminar', 'Estudiante', est.id,
                        valores_anteriores={'nombre': est.nombre, 'rut': est.rut_matricula})
    registrar_cambio_sync('estudiante', est.id, 'actualizar', est)
    flash(f"✅ {est.nombre} dado de baja.")
    return redirect(url_for('ver_inventario'))

# ========== PRESTAMOS ESTUDIANTES ==========

@app.route('/registrar_salida', methods=['POST'])
@login_requerido
@pañolero_o_admin
def registrar_salida():
    estudiante_id = request.form.get('estudiante_id', type=int)
    profesor_id = request.form.get('profesor_id', type=int)
    # Nombre del profesor supervisor (modelo Profesor gestionado por el instructor)
    profesor_nombre = (request.form.get('profesor_nombre') or '').strip() or None
    nombre_practica = (request.form.get('nombre_practica') or '').strip()
    try:
        carrito = json.loads(request.form.get('carrito_data', '[]'))
    except Exception:
        carrito = []
    estudiante = Estudiante.query.get(estudiante_id)
    if not estudiante:
        flash("❌ Estudiante no encontrado.")
        return redirect(url_for('ver_inventario'))
    if not carrito:
        flash("⚠️ No hay items en el carrito.")
        return redirect(url_for('ver_inventario'))

    # Plazo: obligatorio si todos los items son de BIBLIOTECA, opcional en otros.
    plazo_dias = request.form.get('plazo_dias', type=int)

    # Pañolero del día (estudiante designado que atiende este préstamo). Opcional.
    panolero_dia_id = request.form.get('panolero_dia_id', type=int)
    if panolero_dia_id:
        # Validar que sea un pañolero del día activo de esta especialidad
        pd = PanoleroDesignado.query.filter_by(
            estudiante_id=panolero_dia_id,
            especialidad_id=session.get('usuario_especialidad_id'),
            activo=True
        ).first()
        if not pd:
            flash("⚠️ Pañolero del día seleccionado no está activo. El préstamo quedará a nombre del encargado.")
            panolero_dia_id = None

    creados = 0
    ignorados = []  # para diagnóstico visible al usuario
    for entry in carrito:
        # Aceptar tanto 'item_id' (nombre canónico) como 'id' (nombre usado por
        # el carrito del frontend histórico). Sin esto los préstamos caían
        # todos en el continue y "creados" se quedaba en 0.
        item_id = entry.get('item_id') or entry.get('id')
        item = Item.query.get(item_id) if item_id else None
        cantidad = int(entry.get('cantidad') or 1)
        if not item:
            ignorados.append(f"ítem #{item_id or '?'} no encontrado")
            continue
        if item.cantidad_disponible < cantidad:
            ignorados.append(f"{item.nombre}: solicitados {cantidad}, disponibles {item.cantidad_disponible}")
            continue

        # Plazo de devolución: BIBLIOTECA forza 14 días por defecto si no vino otro valor.
        tipo_area_item = (item.especialidad.tipo_area or 'GENERAL') if item.especialidad else 'GENERAL'
        if plazo_dias and plazo_dias > 0:
            fecha_limite = datetime.utcnow() + timedelta(days=plazo_dias)
        elif tipo_area_item == 'BIBLIOTECA':
            fecha_limite = datetime.utcnow() + timedelta(days=14)
        else:
            fecha_limite = None

        item.cantidad_disponible -= cantidad
        # Incrementar usos_actuales si el ítem tiene max_usos configurado (desgaste por uso)
        if item.max_usos:
            item.usos_actuales = (item.usos_actuales or 0) + cantidad

        db.session.add(Prestamo(item_id=item.id, estudiante_id=estudiante.id,
                                profesor_id=profesor_id,
                                profesor_nombre=profesor_nombre,
                                panolero_dia_id=panolero_dia_id,
                                encargado=session.get('usuario_nombre'),
                                cantidad=cantidad, cantidad_solicitada=cantidad,
                                nombre_practica=nombre_practica, estado='Pendiente',
                                fecha_devolucion_esperada=fecha_limite))
        creados += 1
    db.session.commit()
    registrar_auditoria('crear', 'Prestamo', estudiante.id,
                        valores_nuevos={'practica': nombre_practica,
                                        'items_registrados': creados,
                                        'items_ignorados': ignorados[:10]})
    if creados > 0:
        msg = f"✅ {creados} préstamo(s) registrado(s) a {estudiante.nombre}."
    else:
        msg = f"⚠️ No se pudo registrar ningún préstamo a {estudiante.nombre}."
    if ignorados:
        msg += f" Ignorados: {' | '.join(ignorados[:3])}"
        if len(ignorados) > 3:
            msg += f" (y {len(ignorados)-3} más)"
    flash(msg)
    return redirect(url_for('ver_inventario'))

@app.route('/devolver_prestamo/<int:prestamo_id>', methods=['POST'])
@login_requerido
@pañolero_o_admin
def devolver_prestamo(prestamo_id):
    prest = Prestamo.query.get_or_404(prestamo_id)
    cd = int(request.form.get('cantidad_devuelta') or prest.cantidad)
    cm = int(request.form.get('cantidad_mermada') or 0)
    if cd + cm > prest.cantidad:
        flash("❌ Devuelto + mermado supera lo prestado.")
        return redirect(url_for('ver_inventario'))
    prest.item.cantidad_disponible += cd
    if cm > 0:
        prest.item.cantidad_mermada += cm
        prest.item.cantidad_total -= cm
        prest.cantidad_mermada = cm
    prest.estado = 'Devuelto'
    prest.fecha_devolucion = datetime.utcnow()
    db.session.commit()
    registrar_auditoria('actualizar', 'Prestamo', prest.id,
                        valores_nuevos={'devuelto': cd, 'mermado': cm})
    registrar_cambio_sync('prestamo', prest.id, 'actualizar', prest)
    registrar_cambio_sync('item', prest.item.id, 'actualizar', prest.item)
    flash(f"✅ Préstamo #{prest.id} cerrado.")
    return redirect(url_for('ver_inventario'))

# ========== PRESTAMOS EXTERNOS ==========

@app.route('/agregar_prestamo_externo', methods=['POST'])
@login_requerido
@pañolero_o_admin
def agregar_prestamo_externo():
    codigo = (request.form.get('codigo_item') or '').strip()
    cantidad = int(request.form.get('cantidad') or 1)
    esp_dest = (request.form.get('especialidad_destino') or '').strip()
    tipo_p = request.form.get('tipo_persona') or 'externo'
    persona = (request.form.get('persona_retira') or '').strip()
    profesor = (request.form.get('profesor_cargo') or '').strip()
    especialidad_id = session.get('usuario_especialidad_id')
    item = Item.query.filter_by(codigo_barras=codigo, especialidad_id=especialidad_id).first()
    if not item:
        flash(f"❌ Ítem '{codigo}' no existe.")
        return redirect(url_for('ver_inventario'))
    if item.cantidad_disponible < cantidad:
        flash(f"❌ Stock insuficiente: {item.cantidad_disponible}.")
        return redirect(url_for('ver_inventario'))
    # tipo_movimiento: 'prestamo' (se devuelve) o 'consumo' (oficina, no se devuelve)
    tipo_mov = (request.form.get('tipo_movimiento') or 'prestamo').lower().strip()
    if tipo_mov not in ('prestamo', 'consumo'):
        tipo_mov = 'prestamo'

    item.cantidad_disponible -= cantidad
    if tipo_mov == 'consumo':
        # Consumo: descontar también del total (no vuelve nunca)
        item.cantidad_total -= cantidad
    elif item.max_usos:
        # Préstamo (no consumo) cuenta para desgaste si el ítem tiene max_usos
        item.usos_actuales = (item.usos_actuales or 0) + cantidad

    p = PrestamoExterno(item_id=item.id, especialidad_id=especialidad_id,
                        cantidad=cantidad, es_alumno=(tipo_p == 'alumno'),
                        persona_retira=persona, profesor_cargo=profesor,
                        especialidad_destino=esp_dest,
                        encargado=session.get('usuario_nombre'),
                        estado=('Consumido' if tipo_mov == 'consumo' else 'Activo'),
                        tipo_movimiento=tipo_mov,
                        fecha_devolucion=(datetime.utcnow() if tipo_mov == 'consumo' else None))
    db.session.add(p); db.session.commit()
    registrar_auditoria('crear', 'PrestamoExterno', p.id,
                        valores_nuevos={'item': item.nombre, 'persona': persona, 'tipo': tipo_mov})
    accion = 'Consumo registrado' if tipo_mov == 'consumo' else 'Préstamo externo'
    flash(f"✅ {accion}: {item.nombre} → {persona}.")
    return redirect(url_for('ver_inventario'))

@app.route('/devolver_externo/<int:ext_id>', methods=['POST'])
@login_requerido
@pañolero_o_admin
def devolver_externo(ext_id):
    ext = PrestamoExterno.query.get_or_404(ext_id)
    if ext.tipo_movimiento == 'consumo':
        flash("⚠️ Este movimiento es un consumo (oficina), no se puede devolver.")
        return redirect(url_for('ver_inventario'))
    ext.item.cantidad_disponible += ext.cantidad
    ext.estado = 'Devuelto'
    ext.fecha_devolucion = datetime.utcnow()
    db.session.commit()
    registrar_auditoria('actualizar', 'PrestamoExterno', ext.id,
                        valores_nuevos={'estado': 'Devuelto'})
    registrar_cambio_sync('prestamo_externo', ext.id, 'actualizar', ext)
    registrar_cambio_sync('item', ext.item.id, 'actualizar', ext.item)
    flash(f"✅ Préstamo externo #{ext.id} devuelto.")
    return redirect(url_for('ver_inventario'))

# ========== ORDENES DE TRABAJO ==========

@app.route('/agregar_ot', methods=['POST'])
@login_requerido
@pañolero_o_admin
def agregar_ot():
    titulo = (request.form.get('titulo') or '').strip()
    profesional = (request.form.get('profesional_cargo') or '').strip()
    # Alumnos a cargo: pueden venir como checkboxes múltiples (alumnos_cargo[]) o texto libre
    alumnos_lista = request.form.getlist('alumnos_cargo[]')
    if alumnos_lista:
        alumnos = ', '.join(a.strip() for a in alumnos_lista if a.strip())
    else:
        alumnos = (request.form.get('alumnos_cargo') or '').strip()
    descripcion = (request.form.get('descripcion') or '').strip()
    herr = request.form.get('herramientas_utilizadas') or ''
    rep = request.form.get('repuestos_utilizados') or ''
    if not titulo or not profesional:
        flash("❌ Título y profesional obligatorios.")
        return redirect(url_for('ver_inventario'))
    ot = OrdenTrabajo(titulo=titulo, descripcion=descripcion,
                      especialidad_id=session.get('usuario_especialidad_id'),
                      profesional_cargo=profesional, alumnos_cargo=alumnos,
                      herramientas_utilizadas=herr, repuestos_utilizados=rep,
                      profesor_id=session.get('usuario_id'), estado='Pendiente')
    db.session.add(ot); db.session.commit()
    registrar_auditoria('crear', 'OrdenTrabajo', ot.id, valores_nuevos={'titulo': titulo})
    registrar_cambio_sync('orden_trabajo', ot.id, 'crear', ot)
    flash(f"✅ OT #{ot.id} creada.")
    return redirect(url_for('imprimir_ot', ot_id=ot.id))

@app.route('/completar_ot/<int:ot_id>', methods=['POST'])
@login_requerido
@pañolero_o_admin
def completar_ot(ot_id):
    ot = OrdenTrabajo.query.get_or_404(ot_id)
    ot.estado = 'Completada'
    db.session.commit()
    registrar_auditoria('actualizar', 'OrdenTrabajo', ot.id, valores_nuevos={'estado': 'Completada'})
    registrar_cambio_sync('orden_trabajo', ot.id, 'actualizar', ot)
    flash(f"✅ OT #{ot.id} completada.")
    return redirect(url_for('ver_inventario'))

@app.route('/imprimir_ot/<int:ot_id>')
@login_requerido
def imprimir_ot(ot_id):
    ot = OrdenTrabajo.query.get_or_404(ot_id)
    def parse(t):
        if not t: return []
        try:
            d = json.loads(t)
            if isinstance(d, list): return d
        except Exception: pass
        return [{'nombre': l.strip(), 'cantidad': '', 'codigo': ''}
                for l in t.splitlines() if l.strip()]
    return render_template('imprimir_ot.html', ot=ot,
                           herramientas=parse(ot.herramientas_utilizadas),
                           repuestos=parse(ot.repuestos_utilizados))

@app.route('/imprimir_externo/<int:ext_id>')
@login_requerido
def imprimir_externo(ext_id):
    return render_template('imprimir_externo.html',
                           prestamo=PrestamoExterno.query.get_or_404(ext_id))

# ========== HOJA DE VIDA ==========

@app.route('/buscar_hoja_vida', methods=['POST'])
@login_requerido
def buscar_hoja_vida():
    """Acepta código de barras del alumno O RUT/matrícula.
    Si el lector físico escanea el código de barras, busca por esa columna.
    Si el operador digita el RUT, busca por matrícula. Sin distinción.
    """
    raw = (request.form.get('scan_estudiante') or '').strip()
    if not raw:
        flash("⚠️ Debes escanear o digitar el código del alumno.")
        return redirect(url_for('ver_inventario'))

    # 1) Buscar por código de barras (exact match, sin case)
    est = Estudiante.query.filter(
        db.func.lower(Estudiante.codigo_barras) == raw.lower()
    ).first()
    # 2) Si no, buscar por RUT/matrícula
    if not est:
        est = Estudiante.query.filter(
            db.func.lower(Estudiante.rut_matricula) == raw.lower()
        ).first()
    # 3) Si no, búsqueda parcial por RUT (ej: "21345678" cuando el RUT está como "21345678-9")
    if not est:
        like = f"%{raw}%"
        est = Estudiante.query.filter(Estudiante.rut_matricula.ilike(like)).first()

    if not est:
        flash(f"❌ No se encontró ningún alumno con código/RUT: {raw}")
        return redirect(url_for('ver_inventario'))
    return redirect(url_for('hoja_vida', est_id=est.id))

@app.route('/hoja_vida/<int:est_id>')
@login_requerido
def hoja_vida(est_id):
    est = Estudiante.query.get_or_404(est_id)
    prestamos = Prestamo.query.filter_by(estudiante_id=est.id, estado='Pendiente') \
                              .order_by(Prestamo.fecha_prestamo.desc()).all()
    return render_template('hoja_vida.html', estudiante=est, prestamos=prestamos)

@app.route('/procesar_hoja_vida', methods=['POST'])
@login_requerido
@pañolero_o_admin
def procesar_hoja_vida():
    ids = request.form.getlist('prestamo_id[]')
    buenas = request.form.getlist('cant_buena[]')
    malas = request.form.getlist('cant_mala[]')
    procesados = 0
    for pid, cb, cm in zip(ids, buenas, malas):
        try: prest = Prestamo.query.get(int(pid))
        except: continue
        if not prest or prest.estado != 'Pendiente': continue
        cb = int(cb or 0); cm = int(cm or 0)
        if cb + cm == 0: continue
        prest.item.cantidad_disponible += cb
        if cm > 0:
            prest.item.cantidad_mermada += cm
            prest.item.cantidad_total -= cm
            prest.cantidad_mermada = cm
        if cb + cm >= prest.cantidad:
            prest.estado = 'Devuelto'
            prest.fecha_devolucion = datetime.utcnow()
        procesados += 1
    db.session.commit()
    registrar_auditoria('actualizar', 'Prestamo', 0, valores_nuevos={'procesados': procesados})
    flash(f"✅ {procesados} préstamo(s) procesado(s).")
    return redirect(url_for('ver_inventario'))

def _validar_acceso_curso(curso):
    """True si el usuario logueado puede manipular este curso."""
    if session.get('usuario_rol') == 'Admin':
        return True
    return curso.especialidad_id == session.get('usuario_especialidad_id')


@app.route('/curso/editar/<int:curso_id>', methods=['POST'])
@login_requerido
@pañolero_o_admin
def editar_curso(curso_id):
    """Cambia nombre, nivel, letra o año del curso."""
    c = Curso.query.get_or_404(curso_id)
    if not _validar_acceso_curso(c):
        flash("❌ Sin permiso para editar este curso.")
        return redirect(url_for('ver_cursos'))
    nuevo_nombre = (request.form.get('nombre') or '').strip()
    if nuevo_nombre and nuevo_nombre != c.nombre:
        existe = Curso.query.filter_by(nombre=nuevo_nombre,
                                       especialidad_id=c.especialidad_id).first()
        if existe and existe.id != c.id:
            flash(f"⚠️ Ya existe un curso llamado «{nuevo_nombre}» en tu pañol.")
            return redirect(url_for('ver_cursos'))
        c.nombre = nuevo_nombre
    nivel = (request.form.get('nivel') or '').strip()
    if nivel != '':
        c.nivel = nivel or None
    letra = (request.form.get('letra') or '').strip()
    if letra != '':
        c.letra = (letra or None) and letra.upper()[:5]
    anio_raw = (request.form.get('anio') or '').strip()
    if anio_raw:
        try:
            c.anio = int(anio_raw)
        except ValueError:
            pass
    db.session.commit()
    registrar_auditoria('actualizar', 'Curso', c.id,
                        valores_nuevos={'nombre': c.nombre, 'nivel': c.nivel,
                                        'letra': c.letra, 'anio': c.anio})
    flash(f"✅ Curso «{c.nombre}» actualizado.")
    return redirect(url_for('ver_cursos'))


@app.route('/curso/eliminar/<int:curso_id>', methods=['POST'])
@login_requerido
@pañolero_o_admin
def eliminar_curso(curso_id):
    """Elimina un curso (soft-delete).

    Modo por defecto: falla si el curso tiene alumnos activos.
    Modo cascada (form field ``forzar=1``): también da de baja a todos
    los alumnos activos del curso. Los préstamos históricos se conservan
    porque usamos soft-delete (activo=False) en ambos casos.
    """
    c = Curso.query.get_or_404(curso_id)
    if not _validar_acceso_curso(c):
        flash("❌ Sin permiso para eliminar este curso.")
        return redirect(url_for('ver_cursos'))

    forzar = request.form.get('forzar') == '1'
    alumnos = Estudiante.query.filter_by(curso_id=c.id, activo=True).all()
    n_alumnos = len(alumnos)

    if n_alumnos > 0 and not forzar:
        flash(f"⚠️ El curso «{c.nombre}» tiene {n_alumnos} alumno(s). "
              f"Reasigna o elimina los alumnos antes de borrar el curso.")
        return redirect(url_for('ver_cursos'))

    nombre = c.nombre

    # Modo cascada: dar de baja a todos los alumnos activos del curso
    if forzar and alumnos:
        for a in alumnos:
            a.activo = False
            registrar_auditoria('eliminar_cascada', 'Estudiante', a.id,
                                valores_anteriores={'nombre': a.nombre,
                                                    'curso_id': c.id,
                                                    'motivo': f'baja por cascada al eliminar curso «{nombre}»'})

    # Soft delete del curso
    c.activo = False
    db.session.commit()
    registrar_auditoria('eliminar', 'Curso', c.id,
                        valores_anteriores={'nombre': nombre,
                                            'alumnos_dados_de_baja': n_alumnos if forzar else 0,
                                            'modo': 'cascada' if forzar else 'directo'})

    if forzar and n_alumnos > 0:
        flash(f"✅ Curso «{nombre}» eliminado junto con {n_alumnos} alumno(s) (baja en cascada). "
              f"📋 La operación quedó registrada en la auditoría del sistema.")
    else:
        flash(f"✅ Curso «{nombre}» eliminado.")
    return redirect(url_for('ver_cursos'))


@app.route('/eliminar_alumnos_masivo', methods=['POST'])
@login_requerido
@pañolero_o_admin
def eliminar_alumnos_masivo():
    """Da de baja a varios alumnos en una sola operación.

    Recibe del form alguno de estos campos:
      - alumno_ids:        lista de IDs (checkboxes seleccionados)
      - todos_sin_curso=1: borrar TODOS los alumnos sin curso de la especialidad

    Soft delete: marca activo=False, no rompe los préstamos históricos.
    """
    especialidad_id = session.get('usuario_especialidad_id')
    es_admin = session.get('usuario_rol') == 'Admin'
    if not especialidad_id and not es_admin:
        flash("❌ Cuenta sin especialidad.")
        return redirect(url_for('ver_cursos'))

    todos_sin_curso = request.form.get('todos_sin_curso') == '1'
    alumno_ids = request.form.getlist('alumno_ids')

    if todos_sin_curso:
        q = Estudiante.query.filter_by(curso_id=None, activo=True)
        if not es_admin:
            q = q.filter_by(especialidad_id=especialidad_id)
        alumnos = q.all()
    else:
        if not alumno_ids:
            flash("⚠️ No seleccionaste ningún alumno.")
            return redirect(url_for('ver_cursos'))
        try:
            ids = [int(x) for x in alumno_ids]
        except (ValueError, TypeError):
            flash("❌ IDs inválidos.")
            return redirect(url_for('ver_cursos'))
        q = Estudiante.query.filter(Estudiante.id.in_(ids))
        if not es_admin:
            q = q.filter_by(especialidad_id=especialidad_id)
        alumnos = q.all()

    eliminados = 0
    con_deuda = 0
    for est in alumnos:
        # No tocar si el alumno tiene préstamos pendientes
        if Prestamo.query.filter_by(estudiante_id=est.id, estado='Pendiente').first():
            con_deuda += 1
            continue
        est.activo = False
        registrar_cambio_sync('estudiante', est.id, 'actualizar', est)
        eliminados += 1
    db.session.commit()
    registrar_auditoria('eliminar', 'Estudiante', 0,
                        valores_nuevos={'eliminados': eliminados,
                                        'con_deuda_omitidos': con_deuda,
                                        'todos_sin_curso': todos_sin_curso})

    msg = f"✅ {eliminados} alumno(s) dado(s) de baja."
    if con_deuda:
        msg += f" ⚠️ {con_deuda} omitido(s) por tener préstamos pendientes."
    flash(msg)
    return redirect(url_for('ver_cursos'))


@app.route('/curso/asignar_alumnos', methods=['POST'])
@login_requerido
@pañolero_o_admin
def asignar_alumnos_a_curso():
    """Asigna varios alumnos sin curso a un curso destino (creándolo si no existe)."""
    especialidad_id = session.get('usuario_especialidad_id')
    if not especialidad_id and session.get('usuario_rol') != 'Admin':
        flash("❌ Cuenta sin especialidad.")
        return redirect(url_for('ver_cursos'))

    nombre_curso = (request.form.get('curso_destino') or '').strip()
    alumno_ids = request.form.getlist('alumno_ids')
    if not nombre_curso or not alumno_ids:
        flash("❌ Debes elegir un curso y al menos un alumno.")
        return redirect(url_for('ver_cursos'))

    # Get-or-create curso
    c = Curso.query.filter_by(nombre=nombre_curso, especialidad_id=especialidad_id).first()
    if not c:
        import re
        m = re.match(r'^\s*(\d+)\s*[°º]?\s*([A-Za-z])?', nombre_curso)
        nivel = f"{m.group(1)}° Medio" if (m and m.group(1)) else None
        letra = m.group(2).upper() if (m and m.group(2)) else None
        c = Curso(nombre=nombre_curso, nivel=nivel, letra=letra,
                  anio=datetime.now().year,
                  especialidad_id=especialidad_id, activo=True)
        db.session.add(c)
        db.session.flush()

    asignados = 0
    for aid in alumno_ids:
        try:
            est = Estudiante.query.get(int(aid))
        except (ValueError, TypeError):
            continue
        if not est:
            continue
        if session.get('usuario_rol') != 'Admin' and est.especialidad_id != especialidad_id:
            continue
        est.curso_id = c.id
        est.curso = nombre_curso
        asignados += 1
    db.session.commit()
    flash(f"✅ {asignados} alumno(s) asignado(s) al curso «{nombre_curso}».")
    return redirect(url_for('ver_cursos'))


@app.route('/cursos')
@login_requerido
def ver_cursos():
    """Vista de cursos: un panel por curso con sus alumnos.
    El pañolero ve los cursos de SU especialidad. Admin ve todos."""
    if session.get('usuario_rol') == 'Admin':
        cursos = Curso.query.filter_by(activo=True).order_by(Curso.nombre.asc()).all()
        alumnos_sin_curso = Estudiante.query.filter_by(curso_id=None, activo=True).all()
    else:
        especialidad_id = session.get('usuario_especialidad_id')
        cursos = (Curso.query.filter_by(especialidad_id=especialidad_id, activo=True)
                  .order_by(Curso.nombre.asc()).all())
        alumnos_sin_curso = (Estudiante.query
                             .filter_by(especialidad_id=especialidad_id,
                                        curso_id=None, activo=True).all())
    total_alumnos = sum(c.total_alumnos for c in cursos) + len(alumnos_sin_curso)
    return render_template('cursos.html',
                           cursos=cursos,
                           alumnos_sin_curso=alumnos_sin_curso,
                           total_alumnos=total_alumnos)


@app.route('/credenciales')
@login_requerido
def credenciales():
    if session.get('usuario_rol') == 'Admin':
        qs = Estudiante.query.filter_by(activo=True).all()
    else:
        qs = Estudiante.query.filter_by(
            especialidad_id=session.get('usuario_especialidad_id'), activo=True
        ).all()
    estudiantes = [{'rut_matricula': e.rut_matricula, 'nombre': e.nombre,
                    'carrera': e.especialidad.nombre if e.especialidad else '',
                    'curso': e.curso or ''} for e in qs]
    return render_template('credenciales.html', estudiantes=estudiantes)

# ========== EXCEL ==========

@app.route('/cargar_excel', methods=['POST'])
@login_requerido
def cargar_excel():
    """Carga masiva desde Excel.

    Acceso: cualquier usuario autenticado. La operación queda registrada en
    auditoría con usuario_id, rol, nombre del archivo, especialidad destino,
    filas procesadas, ítems creados/actualizados y errores.

    Formato esperado (13 columnas A–M, orden FIJO; las extras son opcionales):
      A (1)  → Código de barra (si está vacía, se autogenera)
      B (2)  → Nombre del ítem (OBLIGATORIO)
      C (3)  → Descripción
      D (4)  → Marca
      E (5)  → Modelo
      F (6)  → Categoría (si está vacía → 'General')
      G (7)  → Cantidad (numérico, OBLIGATORIO)
      H (8)  → Ubicación
      I (9)  → Fecha adquisición (YYYY-MM-DD)
      J (10) → Desgaste ($) (numérico)
      K (11) → Costo unitario ($) (numérico)
      L (12) → Costo total ($) — IGNORADO (lo calcula la app: costo_unit × cantidad)
      M (13) → URL/ruta de imagen de referencia

    Si el código ya existe en esta especialidad, SUMA al stock existente.
    Si es nuevo, lo crea. La primera fila se detecta como cabecera si la
    columna G (Cantidad) no es numérica.
    """
    archivo = request.files.get('archivo_excel')
    if not archivo or archivo.filename == '':
        flash("❌ No subiste ningún archivo.")
        return redirect(url_for('ver_inventario'))

    nombre_archivo = archivo.filename

    LIMITES_ITEM = {
        'codigo_barras': 100, 'nombre': 100, 'marca': 100,
        'modelo': 100, 'categoria': 50, 'ubicacion': 200, 'imagen_url': 500,
    }
    truncados = []
    def tr(valor, campo, idx_fila):
        if valor is None:
            return valor
        limite = LIMITES_ITEM.get(campo)
        if limite and len(str(valor)) > limite:
            truncados.append((idx_fila + 1, campo, len(str(valor))))
            return str(valor)[:limite]
        return valor


    try:
        df = pd.read_excel(archivo, header=None, dtype=object)
    except Exception as e:
        flash(f"❌ No pude leer el Excel: {e}")
        return redirect(url_for('ver_inventario'))

    if len(df) == 0:
        flash("⚠️ El Excel está vacío.")
        return redirect(url_for('ver_inventario'))

    # Detectar si la primera fila es cabecera (columna G = índice 6 = Cantidad, no numérica)
    inicio = 0
    primera = df.iloc[0]
    try:
        int(float(str(primera[6]).strip()))
    except (ValueError, TypeError, IndexError):
        inicio = 1  # primera fila es cabecera, saltarla

    # ── Resolución de la especialidad destino ────────────────────────────
    # Prioridad: (1) la del usuario logueado si tiene; (2) la que venga por
    # formulario (Admin u otros roles sin especialidad fija). NO NULL en BD:
    # si al final queda vacía, se rechaza con mensaje claro (sin caer en 500).
    especialidad_id = session.get('usuario_especialidad_id')
    if not especialidad_id:
        # intentar tomar del formulario (solo Admin u otros roles pueden usarlo)
        try:
            especialidad_id = int(request.form.get('especialidad_id') or 0) or None
        except (TypeError, ValueError):
            especialidad_id = None

    if not especialidad_id:
        flash("❌ Debes seleccionar la especialidad de destino antes de cargar el Excel.")
        return redirect(url_for('ver_inventario'))

    # Validar que la especialidad realmente exista
    if not Especialidad.query.get(especialidad_id):
        flash("❌ La especialidad seleccionada no existe.")
        return redirect(url_for('ver_inventario'))

    def col(fila, i, default=''):
        """Saca el valor de la columna i (0-indexed). Devuelve default si está vacío."""
        try:
            v = fila[i]
        except (IndexError, KeyError):
            return default
        if v is None:
            return default
        if isinstance(v, float) and pd.isna(v):
            return default
        s = str(v).strip()
        if not s or s.lower() == 'nan':
            return default
        return s

    creados = actualizados = 0
    errores = []

    for idx in range(inicio, len(df)):
        fila = df.iloc[idx]
        nombre = col(fila, 1)  # B
        if not nombre:
            continue  # filas vacías al final del Excel

        codigo = col(fila, 0)  # A
        if not codigo:
            # Autogenerar único: timestamp + idx para evitar colisiones
            codigo = datetime.now().strftime('%y%m%d%H%M%S') + f"{idx:04d}"

        descripcion = col(fila, 2, '') or None  # C
        marca = col(fila, 3, '') or None        # D — NUEVO
        modelo = col(fila, 4, '') or None       # E — NUEVO
        categoria = col(fila, 5, 'General')     # F

        try:
            cantidad = int(float(col(fila, 6, '0')))  # G
        except Exception:
            errores.append(f"Fila {idx + 1}: cantidad inválida")
            continue
        if cantidad < 0:
            errores.append(f"Fila {idx + 1}: cantidad negativa")
            continue

        ubicacion = col(fila, 7, '')  # H

        # Fecha adquisición (I) — opcional
        fecha_raw = col(fila, 8, '')
        fecha_adq = None
        if fecha_raw:
            try:
                if hasattr(fecha_raw, 'date'):
                    fecha_adq = fecha_raw.date()
                else:
                    fecha_adq = datetime.fromisoformat(str(fecha_raw)[:10]).date()
            except Exception:
                fecha_adq = None

        # Desgaste $ (J) y costo unitario $ (K) — opcionales
        try:
            desgaste_val = float(col(fila, 9, '0') or 0)
        except Exception:
            desgaste_val = 0.0
        try:
            precio_val = float(col(fila, 10, '0') or 0)
        except Exception:
            precio_val = 0.0

        # Col L (costo total) se ignora: lo calcula la app
        imagen = col(fila, 12, '')  # M

        codigo    = tr(codigo,    'codigo_barras', idx)
        nombre    = tr(nombre,    'nombre',        idx)
        marca     = tr(marca,     'marca',         idx)
        modelo    = tr(modelo,    'modelo',        idx)
        categoria = tr(categoria, 'categoria',     idx)
        ubicacion = tr(ubicacion, 'ubicacion',     idx)
        imagen    = tr(imagen,    'imagen_url',    idx)

        existente = Item.query.filter_by(
            codigo_barras=codigo, especialidad_id=especialidad_id
        ).first()
        if existente:
            existente.cantidad_total += cantidad
            existente.cantidad_disponible += cantidad
            # Solo sobrescribir si el Excel trae valor
            if descripcion:
                existente.descripcion = descripcion
            if marca:
                existente.marca = marca
            if modelo:
                existente.modelo = modelo
            if categoria and categoria != 'General':
                existente.categoria = categoria
            if ubicacion:
                existente.ubicacion = ubicacion
            if imagen:
                existente.imagen_url = imagen
            if fecha_adq:
                existente.fecha_adquisicion = fecha_adq
            if desgaste_val:
                existente.desgaste = desgaste_val
            if precio_val:
                existente.precio_unitario = precio_val
            registrar_cambio_sync('item', existente.id, 'actualizar', existente)
            actualizados += 1
        else:
            nuevo = Item(
                codigo_barras=codigo, nombre=nombre,
                descripcion=descripcion, marca=marca, modelo=modelo,
                categoria=categoria,
                especialidad_id=especialidad_id,
                cantidad_total=cantidad, cantidad_disponible=cantidad,
                ubicacion=ubicacion, imagen_url=imagen,
                fecha_adquisicion=fecha_adq,
                desgaste=desgaste_val, precio_unitario=precio_val,
            )
            db.session.add(nuevo)
            db.session.flush()
            registrar_cambio_sync('item', nuevo.id, 'crear', nuevo)
            creados += 1

    # Commit protegido: si la BD rechaza algo, hacemos rollback y
    # devolvemos mensaje humano en vez de un 500 crudo.
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error al guardar en la base de datos: {str(e)[:180]}")
        return redirect(url_for('ver_inventario'))

    # Auditoría enriquecida: quién, con qué archivo, a qué especialidad,
    # cuántas filas procesó, cuántos creó/actualizó/erró. Queda accesible
    # al Administrador desde /auditoria.
    esp_destino = Especialidad.query.get(especialidad_id)
    registrar_auditoria(
        'importar_excel', 'Item', 0,
        valores_nuevos={
            'archivo': nombre_archivo,
            'usuario_id': session.get('usuario_id'),
            'usuario_nombre': session.get('usuario_nombre'),
            'usuario_rol': session.get('usuario_rol'),
            'especialidad_destino_id': especialidad_id,
            'especialidad_destino_nombre': esp_destino.nombre if esp_destino else None,
            'filas_totales': int(len(df) - inicio),
            'creados': creados,
            'actualizados': actualizados,
            'errores': len(errores),
            'errores_detalle': errores[:10],  # muestra primeros 10 para revisión
            'fecha': datetime.utcnow().isoformat(),
        },
        especialidad_id=especialidad_id
    )

    msg = f"✅ Excel cargado: {creados} ítem(s) nuevo(s), {actualizados} actualizado(s)."
    if errores:
        msg += f" ⚠️ {len(errores)} fila(s) con errores: " + " | ".join(errores[:3])
    msg += " 📋 La carga quedó registrada en la auditoría del sistema."
    flash(msg)
    return redirect(url_for('ver_inventario'))

@app.route('/exportar_excel')
@login_requerido
def exportar_excel():
    if session.get('usuario_rol') == 'Admin':
        items = Item.query.order_by(Item.especialidad_id.asc(),
                                    Item.categoria.asc(),
                                    Item.nombre.asc()).all()
    else:
        items = Item.query.filter_by(
            especialidad_id=session.get('usuario_especialidad_id')
        ).order_by(Item.categoria.asc(), Item.nombre.asc()).all()
    # Formato estándar A–M (13 columnas) — coincide con cargar_excel para roundtrip exacto
    df = pd.DataFrame([{
        'Código de barra': i.codigo_barras,                                      # A
        'Nombre': i.nombre,                                                       # B
        'Descripción': i.descripcion or '',                                       # C
        'Marca': i.marca or '',                                                   # D
        'Modelo': i.modelo or '',                                                 # E
        'Categoría': i.categoria,                                                 # F
        'Cantidad': i.cantidad_total,                                             # G
        'Ubicación': i.ubicacion,                                                 # H
        'Fecha adquisición': i.fecha_adquisicion.isoformat() if i.fecha_adquisicion else '',  # I
        'Desgaste ($)': i.desgaste or 0,                                          # J
        'Costo unitario ($)': i.precio_unitario or 0,                             # K
        'Costo total ($)': (i.precio_unitario or 0) * (i.cantidad_total or 0),    # L
        'Imagen de referencia': i.imagen_url or '',                               # M
        # Columnas auxiliares (informativas; el importador las ignora porque
        # solo lee las primeras 13 columnas en orden)
        'Especialidad': i.especialidad.nombre if i.especialidad else '',
        'Disponible': i.cantidad_disponible,
        'Mermada': i.cantidad_mermada,
    } for i in items])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Inventario', index=False)
    buf.seek(0)
    fname = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ========== USUARIOS ==========

@app.route('/agregar_usuario', methods=['POST'])
@login_requerido
@pañolero_o_admin
def agregar_usuario():
    nombre = (request.form.get('nombre') or '').strip()
    username = (request.form.get('username') or '').strip()
    password = request.form.get('password') or ''
    rol = request.form.get('rol') or 'Profesor'
    if not (nombre and username and password):
        flash("❌ Faltan datos.")
        return redirect(url_for('ver_inventario'))
    if Usuario.query.filter_by(username=username).first():
        flash(f"⚠️ '{username}' ya existe.")
        return redirect(url_for('ver_inventario'))
    u = Usuario(nombre=nombre, username=username,
                password_hash=generate_password_hash(password),
                rol=rol, especialidad_id=session.get('usuario_especialidad_id'), activo=True)
    db.session.add(u); db.session.commit()
    registrar_auditoria('crear', 'Usuario', u.id, valores_nuevos={'username': username, 'rol': rol})
    flash(f"✅ Usuario '{username}' creado.")
    return redirect(url_for('ver_inventario'))

@app.route('/eliminar_usuario/<int:usuario_id>', methods=['POST'])
@login_requerido
@pañolero_o_admin
def eliminar_usuario(usuario_id):
    if usuario_id == session.get('usuario_id'):
        flash("❌ No puedes eliminarte.")
        return redirect(url_for('ver_inventario'))
    u = Usuario.query.get_or_404(usuario_id)
    if u.rol == 'Admin':
        flash("❌ Admin Central no se elimina aquí.")
        return redirect(url_for('ver_inventario'))
    u.activo = False
    db.session.commit()
    registrar_auditoria('eliminar', 'Usuario', u.id, valores_anteriores={'username': u.username})
    flash(f"✅ '{u.username}' desactivado.")
    return redirect(url_for('ver_inventario'))

# ========== GESTION PAÑOLEROS ==========

@app.route('/admin/panoleros')
@login_requerido
@admin_requerido
def gestionar_panoleros():
    """Lista todos los pañoleros (activos e inactivos)."""
    especialidades = Especialidad.query.filter_by(activa=True).order_by(Especialidad.nombre).all()
    panoleros = Usuario.query.filter_by(rol='Pañolero') \
        .order_by(Usuario.activo.desc(), Usuario.nombre.asc()).all()
    return render_template('gestionar_pañoleros.html',
                           especialidades=especialidades,
                           pañoleros=panoleros)

@app.route('/admin/panolero/crear', methods=['POST'])
@login_requerido
@admin_requerido
def crear_panolero():
    """Crea un pañolero. Permite múltiples por especialidad (genera usernames únicos)."""
    import unicodedata
    nombre = (request.form.get('nombre') or '').strip()
    username_custom = (request.form.get('username') or '').strip()
    especialidad_id = request.form.get('especialidad_id', type=int)
    password = (request.form.get('password') or '').strip()
    email_custom = (request.form.get('email') or '').strip()

    if not nombre or not especialidad_id or not password:
        flash("❌ Faltan datos: nombre, especialidad y contraseña son obligatorios.")
        return redirect(url_for('gestionar_panoleros'))

    esp = Especialidad.query.get(especialidad_id)
    if not esp:
        flash("❌ Especialidad no encontrada.")
        return redirect(url_for('gestionar_panoleros'))

    def slugify(s):
        s = ''.join(c for c in unicodedata.normalize('NFD', s)
                    if unicodedata.category(c) != 'Mn')
        return s.lower().replace(' ', '_').replace('.', '_')

    # Usar username custom si viene; si no, generar uno único basado en especialidad
    if username_custom:
        username = slugify(username_custom)
    else:
        base = f"pañolero_{slugify(esp.nombre)}"
        username = base
        n = 2
        while Usuario.query.filter_by(username=username).first():
            username = f"{base}_{n}"
            n += 1

    if Usuario.query.filter_by(username=username).first():
        flash(f"❌ El usuario '{username}' ya existe. Elige otro.")
        return redirect(url_for('gestionar_panoleros'))

    email = email_custom or f"{username}@colegio.local"
    if Usuario.query.filter_by(email=email).first():
        # Email ya en uso, generar único
        email = f"{username}_{datetime.now().strftime('%H%M%S')}@colegio.local"

    nuevo = Usuario(nombre=nombre, username=username, email=email,
                    password_hash=generate_password_hash(password),
                    rol='Pañolero', especialidad_id=especialidad_id, activo=True)
    db.session.add(nuevo); db.session.commit()
    registrar_auditoria('crear', 'Usuario', nuevo.id,
                        valores_nuevos={'nombre': nombre, 'username': username,
                                        'rol': 'Pañolero', 'especialidad': esp.nombre})
    flash(f"✅ Instructor '{nombre}' creado con usuario '{username}'.")
    return redirect(url_for('gestionar_panoleros'))


@app.route('/admin/panolero/<int:panolero_id>/editar', methods=['POST'])
@login_requerido
@admin_requerido
def editar_panolero(panolero_id):
    """Edita datos de un pañolero existente."""
    p = Usuario.query.get(panolero_id)
    if not p or p.rol != 'Pañolero':
        flash("❌ Instructor no encontrado.")
        return redirect(url_for('gestionar_panoleros'))

    nuevo_nombre = (request.form.get('nombre') or '').strip()
    nuevo_username = (request.form.get('username') or '').strip()
    nueva_esp = request.form.get('especialidad_id', type=int)
    nuevo_email = (request.form.get('email') or '').strip()
    nuevo_password = (request.form.get('password') or '').strip()
    activo = request.form.get('activo') == 'on'

    cambios = {}

    if nuevo_nombre and nuevo_nombre != p.nombre:
        cambios['nombre'] = (p.nombre, nuevo_nombre)
        p.nombre = nuevo_nombre

    if nuevo_username and nuevo_username != p.username:
        # Verificar que no choque con otro
        otro = Usuario.query.filter(Usuario.username == nuevo_username, Usuario.id != p.id).first()
        if otro:
            flash(f"❌ Ya existe un usuario con username '{nuevo_username}'.")
            return redirect(url_for('gestionar_panoleros'))
        cambios['username'] = (p.username, nuevo_username)
        p.username = nuevo_username

    if nueva_esp and nueva_esp != p.especialidad_id:
        esp = Especialidad.query.get(nueva_esp)
        if esp:
            cambios['especialidad'] = (
                (p.especialidad_asignada.nombre if p.especialidad_asignada else None),
                esp.nombre
            )
            p.especialidad_id = nueva_esp

    if nuevo_email and nuevo_email != p.email:
        otro = Usuario.query.filter(Usuario.email == nuevo_email, Usuario.id != p.id).first()
        if otro:
            flash(f"❌ Ya existe un usuario con email '{nuevo_email}'.")
            return redirect(url_for('gestionar_panoleros'))
        cambios['email'] = (p.email, nuevo_email)
        p.email = nuevo_email

    if activo != p.activo:
        cambios['activo'] = (p.activo, activo)
        p.activo = activo

    if nuevo_password:
        p.password_hash = generate_password_hash(nuevo_password)
        cambios['password'] = ('***', '***cambiada***')

    if not cambios:
        flash("⚠️ No se detectaron cambios.")
        return redirect(url_for('gestionar_panoleros'))

    db.session.commit()
    registrar_auditoria('actualizar', 'Usuario', p.id, valores_nuevos=cambios)
    flash(f"✅ Instructor '{p.nombre}' actualizado ({len(cambios)} cambio(s)).")
    return redirect(url_for('gestionar_panoleros'))


@app.route('/admin/panolero/<int:panolero_id>/toggle', methods=['POST'])
@login_requerido
@admin_requerido
def toggle_panolero(panolero_id):
    """Activar / desactivar pañolero (sin eliminar)."""
    p = Usuario.query.get(panolero_id)
    if not p or p.rol != 'Pañolero':
        flash("❌ Instructor no encontrado.")
        return redirect(url_for('gestionar_panoleros'))
    p.activo = not p.activo
    db.session.commit()
    estado = 'activado' if p.activo else 'desactivado'
    registrar_auditoria('actualizar', 'Usuario', p.id,
                        valores_nuevos={'estado': estado})
    flash(f"✅ Instructor '{p.nombre}' {estado}.")
    return redirect(url_for('gestionar_panoleros'))

@app.route('/admin/panolero/<int:panolero_id>/eliminar', methods=['POST'])
@login_requerido
@admin_requerido
def eliminar_panolero(panolero_id):
    p = Usuario.query.get(panolero_id)
    if not p or p.rol != 'Pañolero':
        flash("❌ Instructor no encontrado.")
        return redirect(url_for('gestionar_panoleros'))
    registrar_auditoria('eliminar', 'Usuario', panolero_id,
                        valores_anteriores={'nombre': p.nombre, 'username': p.username})
    db.session.delete(p); db.session.commit()
    flash(f"✅ Instructor eliminado.")
    return redirect(url_for('gestionar_panoleros'))

@app.route('/admin/panolero/<int:panolero_id>/resetear-contrasena', methods=['POST'])
@login_requerido
@admin_requerido
def resetear_contrasena_panolero(panolero_id):
    p = Usuario.query.get(panolero_id)
    if not p or p.rol != 'Pañolero':
        flash("❌ Instructor no encontrado.")
        return redirect(url_for('gestionar_panoleros'))
    p.password_hash = generate_password_hash("pañol123")
    db.session.commit()
    registrar_auditoria('actualizar', 'Usuario', panolero_id,
                        valores_nuevos={'accion': 'resetear_contraseña'})
    flash(f"✅ Contraseña de '{p.nombre}' reseteada a: pañol123")
    return redirect(url_for('gestionar_panoleros'))


# ========================================================================
# BIBLIOTECA: catálogo, libros atrasados, alta de libros
# ========================================================================

def _es_biblioteca():
    """True si el usuario actual está en la especialidad Biblioteca o es Admin."""
    rol = session.get('usuario_rol')
    nombre_esp = (session.get('usuario_especialidad') or '').lower()
    return rol == 'Admin' or 'biblioteca' in nombre_esp


@app.route('/biblioteca')
@login_requerido
def biblioteca_catalogo():
    """Catálogo bibliográfico: lista todos los items con datos de libro y permite buscar."""
    q = (request.args.get('q') or '').strip()
    rol = session.get('usuario_rol')

    # Determinar el id de la especialidad Biblioteca
    bib = Especialidad.query.filter_by(nombre='Biblioteca').first()
    if not bib:
        flash("⚠️ La especialidad 'Biblioteca' no existe en la BD.")
        return redirect(url_for('ver_inventario'))

    base_query = Item.query.filter_by(especialidad_id=bib.id)
    if q:
        like = f"%{q}%"
        base_query = base_query.filter(
            db.or_(
                Item.nombre.ilike(like),
                Item.autor.ilike(like),
                Item.isbn.ilike(like),
                Item.editorial.ilike(like),
                Item.codigo_barras.ilike(like),
            )
        )

    libros = base_query.order_by(Item.categoria.asc(), Item.nombre.asc()).limit(500).all()

    # Si la plantilla específica de biblioteca no existe aún, reutilizamos inventario.html
    # con un set de variables compatible.
    return render_template('inventario.html',
                           items=libros,
                           estudiantes=Estudiante.query.filter_by(activo=True).all(),
                           prestamos=Prestamo.query.join(Item).filter(
                               Item.especialidad_id == bib.id
                           ).order_by(Prestamo.fecha_prestamo.desc()).limit(100).all(),
                           prestamos_externos=[],
                           ordenes_trabajo=[],
                           profesores=Usuario.query.filter(Usuario.rol == 'Profesor').all(),
                           usuarios_sistema=[],
                           practicas_resumen=[],
                           total_stock=sum(l.cantidad_total for l in libros),
                           prestamos_activos=Prestamo.query.join(Item).filter(
                               Item.especialidad_id == bib.id,
                               Prestamo.estado == 'Pendiente'
                           ).count(),
                           alertas=[],
                           especialidad='Biblioteca')


@app.route('/biblioteca/atrasados')
@login_requerido
def biblioteca_atrasados():
    """Devuelve JSON con los préstamos atrasados (días > 0)."""
    bib = Especialidad.query.filter_by(nombre='Biblioteca').first()
    if not bib:
        return jsonify({'atrasados': [], 'total': 0})

    pendientes = Prestamo.query.join(Item).filter(
        Item.especialidad_id == bib.id,
        Prestamo.estado == 'Pendiente',
        Prestamo.fecha_devolucion_esperada.isnot(None)
    ).all()

    atrasados = []
    for p in pendientes:
        if p.dias_atraso > 0:
            atrasados.append({
                'prestamo_id': p.id,
                'libro': p.item.nombre,
                'autor': p.item.autor or '',
                'isbn': p.item.isbn or '',
                'estudiante': p.estudiante.nombre if p.estudiante else '',
                'curso': p.estudiante.curso if p.estudiante else '',
                'fecha_prestamo': p.fecha_prestamo.strftime('%Y-%m-%d'),
                'fecha_limite': p.fecha_devolucion_esperada.strftime('%Y-%m-%d'),
                'dias_atraso': p.dias_atraso,
            })
    return jsonify({'atrasados': atrasados, 'total': len(atrasados)})


@app.route('/agregar_libro', methods=['POST'])
@login_requerido
@pañolero_o_admin
def agregar_libro():
    """Alta de libro con campos bibliográficos completos."""
    bib = Especialidad.query.filter_by(nombre='Biblioteca').first()
    if not bib:
        flash("❌ La especialidad Biblioteca no existe.")
        return redirect(url_for('ver_inventario'))

    titulo = (request.form.get('nombre') or '').strip()
    if not titulo:
        flash("❌ Falta el título del libro.")
        return redirect(url_for('biblioteca_catalogo'))

    isbn = (request.form.get('isbn') or '').strip() or None
    autor = (request.form.get('autor') or '').strip() or None
    editorial = (request.form.get('editorial') or '').strip() or None
    try:
        anio = int(request.form.get('anio_publicacion') or 0) or None
    except Exception:
        anio = None
    cantidad = int(request.form.get('cantidad') or 1)
    ubicacion = (request.form.get('ubicacion') or 'Sin especificar').strip()

    # Si el ISBN existe en biblioteca, sumar stock
    existente = None
    if isbn:
        existente = Item.query.filter_by(especialidad_id=bib.id, isbn=isbn).first()
    if existente:
        existente.cantidad_total += cantidad
        existente.cantidad_disponible += cantidad
        item_id = existente.id
        accion = 'actualizar'
    else:
        codigo = isbn or datetime.now().strftime('%y%m%d%H%M%S') + "L"
        nuevo = Item(
            codigo_barras=codigo, nombre=titulo, categoria='Biblioteca',
            especialidad_id=bib.id,
            cantidad_total=cantidad, cantidad_disponible=cantidad,
            ubicacion=ubicacion,
            autor=autor, isbn=isbn, editorial=editorial, anio_publicacion=anio,
        )
        db.session.add(nuevo)
        db.session.flush()
        item_id = nuevo.id
        accion = 'crear'
    db.session.commit()
    registrar_auditoria(accion, 'Item', item_id,
                        valores_nuevos={'titulo': titulo, 'autor': autor, 'isbn': isbn})
    flash(f"✅ Libro '{titulo}' registrado.")
    return redirect(url_for('biblioteca_catalogo'))


# ========================================================================
# OFICINA: registro de consumos (sin devolución)
# ========================================================================

@app.route('/registrar_consumo', methods=['POST'])
@login_requerido
@pañolero_o_admin
def registrar_consumo():
    """Registra un consumo de oficina. Descuenta del stock TOTAL (no vuelve)."""
    codigo = (request.form.get('codigo_item') or '').strip()
    cantidad = int(request.form.get('cantidad') or 1)
    persona = (request.form.get('persona_retira') or '').strip()
    motivo = (request.form.get('motivo') or '').strip()

    especialidad_id = session.get('usuario_especialidad_id')
    item = Item.query.filter_by(codigo_barras=codigo, especialidad_id=especialidad_id).first()
    if not item:
        flash(f"❌ Ítem '{codigo}' no existe en este pañol/área.")
        return redirect(url_for('ver_inventario'))
    if item.cantidad_disponible < cantidad:
        flash(f"❌ Stock insuficiente: {item.cantidad_disponible}.")
        return redirect(url_for('ver_inventario'))
    if not persona:
        flash("❌ Indica quién retira el consumible.")
        return redirect(url_for('ver_inventario'))

    item.cantidad_disponible -= cantidad
    item.cantidad_total -= cantidad

    consumo = PrestamoExterno(
        item_id=item.id, especialidad_id=especialidad_id,
        cantidad=cantidad, es_alumno=False,
        persona_retira=persona, profesor_cargo='',
        especialidad_destino=motivo or 'Consumo interno',
        encargado=session.get('usuario_nombre'),
        estado='Consumido', tipo_movimiento='consumo',
        fecha_devolucion=datetime.utcnow()
    )
    db.session.add(consumo)
    db.session.commit()
    registrar_auditoria('consumir', 'Item', item.id,
                        valores_nuevos={'item': item.nombre, 'cantidad': cantidad, 'persona': persona})
    flash(f"✅ Consumo registrado: {cantidad} × {item.nombre} → {persona}.")
    return redirect(url_for('ver_inventario'))


# ========================================================================
# API DE SINCRONIZACIÓN (solo cuando ES_ADMIN_CENTRAL)
# ========================================================================

def _verificar_token_sync():
    """Lee el header Authorization y valida contra PANOL_SYNC_TOKEN."""
    auth = request.headers.get('Authorization', '')
    if not auth.startswith('Bearer '):
        return False
    return auth[7:].strip() == PANOL_SYNC_TOKEN


@app.route('/api/sync/push', methods=['POST'])
def api_sync_push():
    """Recibe un lote de cambios desde un nodo. Solo el admin central acepta esto.

    Body JSON: {
      "nodo": "panol_electronica",
      "cambios": [
        {"sync_uuid": "...", "tabla": "item", "registro_id_local": 12,
         "accion": "crear", "payload": {...}},
        ...
      ]
    }
    Respuesta: {"ok": [uuid,...], "error": {uuid: msg, ...}}
    """
    if not ES_ADMIN_CENTRAL:
        return jsonify({'error': 'Este nodo no es admin central'}), 403
    if not _verificar_token_sync():
        return jsonify({'error': 'Token inválido'}), 401

    try:
        data = request.get_json(force=True)
    except Exception:
        return jsonify({'error': 'JSON inválido'}), 400

    nodo = (data or {}).get('nodo', 'desconocido')
    cambios = (data or {}).get('cambios', [])

    ok_uuids = []
    err_map = {}

    for ch in cambios:
        uid = ch.get('sync_uuid')
        try:
            _aplicar_cambio_remoto(nodo, ch)
            ok_uuids.append(uid)
        except Exception as e:
            err_map[uid] = f"{type(e).__name__}: {e}"
            db.session.rollback()

    db.session.commit()
    return jsonify({'ok': ok_uuids, 'error': err_map,
                    'recibidos': len(cambios), 'aplicados': len(ok_uuids)})


def _aplicar_cambio_remoto(nodo, ch):
    """Aplica un cambio recibido desde un nodo en la BD del admin central.

    Estrategia: localiza el registro por (nodo + registro_id_local) usando código de barras
    o RUT como clave natural cuando aplica. Si no existe, lo crea. Si existe, actualiza.
    """
    tabla = ch['tabla']
    accion = ch['accion']
    payload = ch.get('payload') or {}

    # Idempotencia: si ya recibimos este sync_uuid, ignorar
    uid = ch.get('sync_uuid')
    if uid:
        existente_uuid = SyncLog.query.filter_by(sync_uuid=uid).first()
        if existente_uuid and existente_uuid.push_status == 'aplicado':
            return  # ya lo aplicamos antes
        if not existente_uuid:
            db.session.add(SyncLog(
                nodo_origen=nodo, tabla=tabla,
                registro_id_local=ch.get('registro_id_local', 0),
                accion=accion,
                payload=json.dumps(payload),
                sync_uuid=uid,
                push_status='aplicado',
            ))

    if tabla == 'item':
        _upsert_item(nodo, payload, accion)
    elif tabla == 'estudiante':
        _upsert_estudiante(nodo, payload, accion)
    elif tabla == 'prestamo':
        _upsert_prestamo(nodo, payload, accion)
    elif tabla == 'prestamo_externo':
        _upsert_prestamo_externo(nodo, payload, accion)
    elif tabla == 'orden_trabajo':
        _upsert_orden_trabajo(nodo, payload, accion)
    else:
        raise ValueError(f"Tabla {tabla} no soportada")


def _to_dt(s):
    if not s:
        return None
    try:
        return datetime.fromisoformat(s)
    except Exception:
        return None


def _upsert_item(nodo, p, accion):
    """Localiza item por código de barras + especialidad_id, o crea."""
    codigo = p.get('codigo_barras')
    esp_id = p.get('especialidad_id')
    item = None
    if codigo and esp_id:
        item = Item.query.filter_by(codigo_barras=codigo, especialidad_id=esp_id).first()
    if accion == 'eliminar':
        if item:
            db.session.delete(item)
        return
    if not item:
        item = Item(codigo_barras=codigo, especialidad_id=esp_id, nombre=p.get('nombre', ''))
        db.session.add(item)
    # Copiar campos seguros (no id)
    for k in ('nombre', 'descripcion', 'categoria', 'cantidad_total',
              'cantidad_disponible', 'cantidad_mermada', 'cantidad_minima',
              'imagen_url', 'ubicacion', 'precio_unitario', 'desgaste',
              'autor', 'isbn', 'editorial', 'anio_publicacion',
              'marca', 'modelo', 'numero_serie', 'estado',
              'fecha_adquisicion', 'max_usos'):
        if k in p:
            setattr(item, k, p[k])


def _upsert_estudiante(nodo, p, accion):
    rut = p.get('rut_matricula')
    if not rut:
        raise ValueError("Estudiante sin RUT")
    est = Estudiante.query.filter_by(rut_matricula=rut).first()
    if not est:
        est = Estudiante(rut_matricula=rut, nombre=p.get('nombre', ''),
                         especialidad_id=p.get('especialidad_id'),
                         password_hash=p.get('password_hash', generate_password_hash(rut)))
        db.session.add(est)
    for k in ('nombre', 'curso', 'especialidad_id', 'activo', 'password_hash'):
        if k in p:
            setattr(est, k, p[k])


def _upsert_prestamo(nodo, p, accion):
    """Para préstamos: usamos clave compuesta (nodo, registro_id_local) almacenada como prefijo."""
    # Mapeamos local id a un registro en el admin con descripción única
    encargado_marker = f"[{nodo}#{p.get('id')}] " + (p.get('encargado') or '')
    pres = Prestamo.query.filter(Prestamo.encargado.like(f"[{nodo}#{p.get('id')}]%")).first()

    item = None
    if p.get('item_id'):
        # En el admin no podemos usar el item_id local; tenemos que resolver por código.
        # Asumimos que el item ya fue sincronizado antes. Si no existe, fallamos suave.
        # (Una mejora futura: incluir codigo_barras en el payload del préstamo.)
        item = Item.query.get(p.get('item_id'))
    if not item:
        # Sin item válido no podemos persistir un préstamo nuevo
        if not pres:
            raise ValueError(f"Item local {p.get('item_id')} no encontrado en admin")

    if not pres:
        pres = Prestamo(item_id=item.id if item else 0,
                        estudiante_id=p.get('estudiante_id') or 0,
                        cantidad=p.get('cantidad') or 1,
                        encargado=encargado_marker)
        db.session.add(pres)

    for k in ('cantidad', 'cantidad_solicitada', 'cantidad_mermada',
              'nombre_practica', 'estado', 'multa'):
        if k in p:
            setattr(pres, k, p[k])
    pres.fecha_prestamo = _to_dt(p.get('fecha_prestamo')) or pres.fecha_prestamo
    pres.fecha_devolucion = _to_dt(p.get('fecha_devolucion'))
    pres.fecha_devolucion_esperada = _to_dt(p.get('fecha_devolucion_esperada'))



@app.route('/descargar_plantilla')
@login_requerido
def descargar_plantilla():
    """Descarga el archivo inventario_muestra.xlsx como plantilla."""
    aqui = os.path.dirname(os.path.abspath(__file__))
    plantilla = os.path.join(aqui, 'inventario_muestra.xlsx')
    if os.path.exists(plantilla):
        return send_file(plantilla, as_attachment=True,
                         download_name='plantilla_inventario.xlsx')
    # Fallback: generar al vuelo con las 11 columnas
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'
    headers = ['Codigo de barra', 'Nombre', 'Descripcion', 'Categoria',
               'Cantidad', 'Ubicacion', 'Fecha adquisicion',
               'Desgaste ($)', 'Costo unitario ($)', 'Costo total ($)',
               'Imagen de referencia']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='plantilla_inventario.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/descargar_plantilla_alumnos')
@login_requerido
def descargar_plantilla_alumnos():
    """Descarga la plantilla Excel para carga masiva de alumnos (2 columnas)."""
    aqui = os.path.dirname(os.path.abspath(__file__))
    plantilla = os.path.join(aqui, 'alumnos_muestra.xlsx')
    if os.path.exists(plantilla):
        return send_file(plantilla, as_attachment=True,
                         download_name='plantilla_alumnos.xlsx')
    # Fallback: 2 columnas
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Alumnos'
    ws.cell(row=1, column=1, value='N de lista')
    ws.cell(row=1, column=2, value='Nombre completo')
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='plantilla_alumnos.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/reporte_mermas')
@login_requerido
@admin_requerido
def admin_reporte_mermas():
    """Reporte de mermas Excel con columnas adaptadas al tipo de cada área.
    Una hoja por especialidad. Solo ítems con cantidad_mermada > 0 + préstamos
    cerrados con merma. PANOL_TP muestra ubicación/categoría; BIBLIOTECA muestra
    ISBN/autor; INFORMATICA muestra marca/modelo/serie; DEPORTIVO muestra estado."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        creadas = 0
        for esp in Especialidad.query.filter_by(activa=True).order_by(Especialidad.id).all():
            items_merma = Item.query.filter(
                Item.especialidad_id == esp.id,
                Item.cantidad_mermada > 0
            ).order_by(Item.cantidad_mermada.desc()).all()
            if not items_merma:
                continue
            ta = esp.tipo_area or 'GENERAL'

            # Columnas por tipo
            if ta == 'BIBLIOTECA':
                rows = [{
                    'ISBN':        i.isbn or '',
                    'Título':      i.nombre,
                    'Autor':       i.autor or '',
                    'Editorial':   i.editorial or '',
                    'Categoría':   i.categoria or '',
                    'Mermados':    i.cantidad_mermada,
                    'Disponibles': i.cantidad_disponible,
                    'Total':       i.cantidad_total,
                    'Costo unit ($)':  i.precio_unitario or 0,
                    'Pérdida ($)':     (i.precio_unitario or 0) * (i.cantidad_mermada or 0),
                    'Ubicación':   i.ubicacion or '',
                } for i in items_merma]
            elif ta == 'INFORMATICA':
                rows = [{
                    'Código':      i.codigo_barras,
                    'Nombre':      i.nombre,
                    'Marca':       i.marca or '',
                    'Modelo':      i.modelo or '',
                    'N° Serie':    i.numero_serie or '',
                    'Estado':      i.estado or '',
                    'Categoría':   i.categoria or '',
                    'Mermados':    i.cantidad_mermada,
                    'Disponibles': i.cantidad_disponible,
                    'Total':       i.cantidad_total,
                    'Costo unit ($)':  i.precio_unitario or 0,
                    'Pérdida ($)':     (i.precio_unitario or 0) * (i.cantidad_mermada or 0),
                    'Ubicación':   i.ubicacion or '',
                } for i in items_merma]
            elif ta == 'DEPORTIVO':
                rows = [{
                    'Código':      i.codigo_barras,
                    'Nombre':      i.nombre,
                    'Categoría':   i.categoria or '',
                    'Estado':      i.estado or '',
                    'Mermados':    i.cantidad_mermada,
                    'Disponibles': i.cantidad_disponible,
                    'Total':       i.cantidad_total,
                    'Costo unit ($)':  i.precio_unitario or 0,
                    'Pérdida ($)':     (i.precio_unitario or 0) * (i.cantidad_mermada or 0),
                    'Ubicación':   i.ubicacion or '',
                } for i in items_merma]
            elif ta == 'PANOL_TP':
                rows = [{
                    'Código':      i.codigo_barras,
                    'Nombre':      i.nombre,
                    'Marca':       i.marca or '',
                    'Modelo':      i.modelo or '',
                    'Categoría':   i.categoria or '',
                    'Mermados':    i.cantidad_mermada,
                    'Disponibles': i.cantidad_disponible,
                    'Total':       i.cantidad_total,
                    'Costo unit ($)':  i.precio_unitario or 0,
                    'Pérdida ($)':     (i.precio_unitario or 0) * (i.cantidad_mermada or 0),
                    'Ubicación':   i.ubicacion or '',
                } for i in items_merma]
            else:  # GENERAL
                rows = [{
                    'Código':      i.codigo_barras,
                    'Nombre':      i.nombre,
                    'Categoría':   i.categoria or '',
                    'Mermados':    i.cantidad_mermada,
                    'Disponibles': i.cantidad_disponible,
                    'Total':       i.cantidad_total,
                    'Costo unit ($)':  i.precio_unitario or 0,
                    'Pérdida ($)':     (i.precio_unitario or 0) * (i.cantidad_mermada or 0),
                    'Ubicación':   i.ubicacion or '',
                } for i in items_merma]

            df = pd.DataFrame(rows)
            # Nombre de hoja máx 31 chars en Excel
            sheet = (esp.nombre[:28] + '...') if len(esp.nombre) > 31 else esp.nombre
            df.to_excel(writer, sheet_name=sheet, index=False)
            creadas += 1

        if creadas == 0:
            # Hoja vacía si no hay mermas
            pd.DataFrame([{'Aviso': 'No hay items con merma en ningún área.'}])                 .to_excel(writer, sheet_name='Sin mermas', index=False)

    buf.seek(0)
    fname = f"reporte_mermas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/exportar_completo')
@login_requerido
@admin_requerido
def admin_exportar_completo():
    """Exporta TODO el inventario consolidado por area."""
    items = Item.query.order_by(Item.especialidad_id.asc(),
                                Item.categoria.asc(), Item.nombre.asc()).all()
    df = pd.DataFrame([{
        'Especialidad':   i.especialidad.nombre if i.especialidad else '',
        'Codigo':         i.codigo_barras,
        'Nombre':         i.nombre,
        'Categoria':      i.categoria,
        'Cantidad total': i.cantidad_total,
        'Disponible':     i.cantidad_disponible,
        'Mermada':        i.cantidad_mermada,
        'Ubicacion':      i.ubicacion,
        'Costo unitario': i.precio_unitario or 0,
        'Desgaste':       i.desgaste if hasattr(i, 'desgaste') else 0,
        'Costo total':    (i.precio_unitario or 0) * (i.cantidad_total or 0),
    } for i in items])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Consolidado', index=False)
    buf.seek(0)
    fname = f"inventario_consolidado_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



@app.route('/exportar_inventario')
@login_requerido
def exportar_inventario():
    """Alias de /exportar_excel para mantener compatibilidad con la plantilla."""
    return exportar_excel()


# ========================================================================
# CURSOS A CARGO — máx. 2 cursos "propios" por pañol
# ========================================================================

def _cursos_a_cargo(especialidad_id):
    """Lista de cursos marcados como a_cargo=True para una especialidad (máx. 2)."""
    if not especialidad_id:
        return []
    try:
        return Curso.query.filter_by(
            especialidad_id=especialidad_id, activo=True, a_cargo=True
        ).order_by(Curso.nombre.asc()).all()
    except Exception as e:
        print(f"[WARN] _cursos_a_cargo: {e}")
        return []


def _alumnos_visitantes(especialidad_id):
    """Alumnos de OTRAS especialidades que han hecho préstamos en este pañol.
    Devuelve lista de dicts: {'estudiante': Estudiante, 'num_prestamos', 'ultima_fecha'}.
    """
    if not especialidad_id:
        return []
    try:
        rows = db.session.query(
            Estudiante,
            db.func.count(Prestamo.id).label('num_prestamos'),
            db.func.max(Prestamo.fecha_prestamo).label('ultima_fecha')
        ).join(Prestamo, Prestamo.estudiante_id == Estudiante.id) \
         .join(Item, Item.id == Prestamo.item_id) \
         .filter(Item.especialidad_id == especialidad_id,
                 Estudiante.especialidad_id != especialidad_id) \
         .group_by(Estudiante.id) \
         .order_by(db.func.count(Prestamo.id).desc(),
                   db.func.max(Prestamo.fecha_prestamo).desc()) \
         .limit(50).all()
        return [{'estudiante': r[0], 'num_prestamos': r[1], 'ultima_fecha': r[2]} for r in rows]
    except Exception as e:
        print(f"[WARN] _alumnos_visitantes: {e}")
        return []


@app.route('/pañol/curso/marcar_a_cargo/<int:curso_id>', methods=['POST'])
@login_requerido
@pañolero_o_admin
def marcar_curso_a_cargo(curso_id):
    """Marca un curso como 'a cargo' del pañol. Máx. MAX_CURSOS_A_CARGO activos por especialidad."""
    especialidad_id = session.get('usuario_especialidad_id')
    if not especialidad_id:
        flash("❌ No tienes una especialidad activa.")
        return redirect(url_for('ver_inventario'))
    curso = Curso.query.get_or_404(curso_id)
    if curso.especialidad_id != especialidad_id and session.get('usuario_rol') != 'Admin':
        flash("❌ Ese curso no pertenece a tu especialidad.")
        return redirect(url_for('ver_inventario'))
    if curso.a_cargo:
        flash(f"ℹ️ '{curso.nombre}' ya está marcado como a cargo.")
        return redirect(url_for('ver_inventario'))
    actuales = _cursos_a_cargo(especialidad_id)
    if len(actuales) >= MAX_CURSOS_A_CARGO:
        nombres = ', '.join(c.nombre for c in actuales)
        flash(f"⚠️ Ya tienes {MAX_CURSOS_A_CARGO} cursos a cargo ({nombres}). Quita uno antes.")
        return redirect(url_for('ver_inventario'))
    curso.a_cargo = True
    db.session.commit()
    try:
        registrar_auditoria('actualizar', 'Curso', curso.id,
                            valores_nuevos={'a_cargo': True, 'curso': curso.nombre})
    except Exception:
        pass
    flash(f"✅ '{curso.nombre}' marcado como curso a cargo.")
    return redirect(url_for('ver_inventario'))


@app.route('/pañol/curso/quitar_a_cargo/<int:curso_id>', methods=['POST'])
@login_requerido
@pañolero_o_admin
def quitar_curso_a_cargo(curso_id):
    """Quita el flag 'a cargo' de un curso. Los alumnos no se borran."""
    especialidad_id = session.get('usuario_especialidad_id')
    curso = Curso.query.get_or_404(curso_id)
    if curso.especialidad_id != especialidad_id and session.get('usuario_rol') != 'Admin':
        flash("❌ Sin permiso.")
        return redirect(url_for('ver_inventario'))
    if not curso.a_cargo:
        flash(f"ℹ️ '{curso.nombre}' no estaba marcado como a cargo.")
        return redirect(url_for('ver_inventario'))
    curso.a_cargo = False
    db.session.commit()
    try:
        registrar_auditoria('actualizar', 'Curso', curso.id,
                            valores_nuevos={'a_cargo': False, 'curso': curso.nombre})
    except Exception:
        pass
    flash(f"✅ '{curso.nombre}' dado de baja como curso a cargo (los alumnos siguen en la BD).")
    return redirect(url_for('ver_inventario'))


# ========================================================================
# PROFESORES SUPERVISORES — gestionados por el instructor de cada área
# ========================================================================

def _profesores_activos(especialidad_id):
    """Lista de profesores activos de una especialidad (orden alfabético)."""
    if not especialidad_id:
        return []
    try:
        return Profesor.query.filter_by(
            especialidad_id=especialidad_id, activo=True
        ).order_by(Profesor.nombre.asc()).all()
    except Exception as e:
        print(f"[WARN] _profesores_activos: {e}")
        return []


@app.route('/profesor/agregar', methods=['POST'])
@login_requerido
@pañolero_o_admin
def agregar_profesor():
    """El instructor registra un profesor supervisor de su área."""
    especialidad_id = session.get('usuario_especialidad_id')
    if not especialidad_id:
        flash("❌ No tienes una especialidad activa.")
        return redirect(url_for('ver_inventario'))
    nombre = (request.form.get('nombre') or '').strip()
    if not nombre:
        flash("❌ Debes escribir el nombre del profesor.")
        return redirect(url_for('ver_inventario'))
    # Evitar duplicados activos
    existe = Profesor.query.filter(
        Profesor.especialidad_id == especialidad_id,
        Profesor.activo == True,
        db.func.lower(Profesor.nombre) == nombre.lower()
    ).first()
    if existe:
        flash(f"ℹ️ El profesor '{nombre}' ya está registrado.")
        return redirect(url_for('ver_inventario'))
    db.session.add(Profesor(nombre=nombre, especialidad_id=especialidad_id))
    db.session.commit()
    try:
        registrar_auditoria('crear', 'Profesor', 0, valores_nuevos={'nombre': nombre})
    except Exception:
        pass
    flash(f"✅ Profesor '{nombre}' registrado.")
    return redirect(url_for('ver_inventario'))


@app.route('/profesor/quitar/<int:profesor_id>', methods=['POST'])
@login_requerido
@pañolero_o_admin
def quitar_profesor(profesor_id):
    """Da de baja un profesor supervisor."""
    especialidad_id = session.get('usuario_especialidad_id')
    p = Profesor.query.get_or_404(profesor_id)
    if p.especialidad_id != especialidad_id and session.get('usuario_rol') != 'Admin':
        flash("❌ Sin permiso.")
        return redirect(url_for('ver_inventario'))
    p.activo = False
    db.session.commit()
    flash(f"✅ Profesor '{p.nombre}' dado de baja.")
    return redirect(url_for('ver_inventario'))


@app.route('/panoleros_dia/agregar', methods=['POST'])
@login_requerido
@pañolero_o_admin
def panoleros_dia_agregar():
    """Designa un estudiante como panolero del dia. Maximo MAX_PANOLEROS_DIA por especialidad."""
    especialidad_id = session.get('usuario_especialidad_id')
    if not especialidad_id:
        flash("Cuenta sin especialidad.")
        return redirect(url_for('ver_inventario'))
    estudiante_id = request.form.get('estudiante_id', type=int)
    if not estudiante_id:
        flash("Selecciona un estudiante.")
        return redirect(url_for('ver_inventario'))
    activos = PanoleroDesignado.query.filter_by(
        especialidad_id=especialidad_id, activo=True).count()
    if activos >= MAX_PANOLEROS_DIA:
        flash(f"Ya hay {MAX_PANOLEROS_DIA} panoleros del dia activos. Quita alguno antes.")
        return redirect(url_for('ver_inventario'))
    existe = PanoleroDesignado.query.filter_by(
        estudiante_id=estudiante_id, especialidad_id=especialidad_id, activo=True).first()
    if existe:
        flash("Ese estudiante ya esta designado.")
        return redirect(url_for('ver_inventario'))
    nuevo = PanoleroDesignado(
        estudiante_id=estudiante_id,
        especialidad_id=especialidad_id,
        designado_por_id=session.get('usuario_id'),
        activo=True)
    db.session.add(nuevo)
    db.session.commit()
    flash("Panolero del dia agregado.")
    return redirect(url_for('ver_inventario'))


@app.route('/panoleros_dia/quitar/<int:pd_id>', methods=['POST'])
@login_requerido
@pañolero_o_admin
def panoleros_dia_quitar(pd_id):
    """Da de baja a un panolero del dia."""
    pd = PanoleroDesignado.query.get_or_404(pd_id)
    if session.get('usuario_rol') != 'Admin' and pd.especialidad_id != session.get('usuario_especialidad_id'):
        flash("Sin permiso.")
        return redirect(url_for('ver_inventario'))
    pd.activo = False
    pd.fecha_baja = datetime.utcnow()
    db.session.commit()
    flash("Panolero del dia quitado.")
    return redirect(url_for('ver_inventario'))


@app.route('/panoleros_dia/limpiar', methods=['POST'])
@login_requerido
@pañolero_o_admin
def panoleros_dia_limpiar():
    """Da de baja a TODOS los panoleros del dia de la especialidad."""
    especialidad_id = session.get('usuario_especialidad_id')
    if not especialidad_id:
        flash("Cuenta sin especialidad.")
        return redirect(url_for('ver_inventario'))
    n = 0
    for pd in PanoleroDesignado.query.filter_by(
            especialidad_id=especialidad_id, activo=True).all():
        pd.activo = False
        pd.fecha_baja = datetime.utcnow()
        n += 1
    db.session.commit()
    flash(f"{n} panolero(s) del dia removido(s).")
    return redirect(url_for('ver_inventario'))


@app.route('/admin/buscar')
@login_requerido
@admin_requerido
def admin_buscar():
    """Busqueda global de items para admin."""
    q = (request.args.get('q') or '').strip()
    if not q:
        return render_template('admin_buscar.html', items=[], q='')
    like = f"%{q}%"
    items = Item.query.filter(
        db.or_(
            Item.nombre.ilike(like),
            Item.codigo_barras.ilike(like),
            Item.autor.ilike(like) if hasattr(Item, 'autor') else False,
            Item.isbn.ilike(like) if hasattr(Item, 'isbn') else False,
        )
    ).order_by(Item.especialidad_id.asc(), Item.nombre.asc()).limit(200).all()
    return render_template('admin_buscar.html', items=items, q=q)


@app.route('/admin/cambiar_password', methods=['GET', 'POST'])
@login_requerido
def admin_cambiar_password():
    """Cambiar contraseña del usuario logueado. Los nombres de los campos del form
    DEBEN coincidir con el template cambiar_password.html: actual / nueva / confirmar."""
    user = Usuario.query.get(session.get('usuario_id'))
    if not user:
        flash("❌ Sesión inválida.")
        return redirect(url_for('login'))
    if request.method == 'POST':
        actual = request.form.get('actual', '')
        nueva = request.form.get('nueva', '')
        confirmar = request.form.get('confirmar', '')

        if not check_password_hash(user.password_hash, actual):
            flash("❌ La contraseña actual es incorrecta.")
            return redirect(url_for('admin_cambiar_password'))
        if not nueva or len(nueva) < 8:
            flash("❌ La nueva contraseña debe tener al menos 8 caracteres.")
            return redirect(url_for('admin_cambiar_password'))
        if nueva != confirmar:
            flash("❌ La confirmación no coincide con la nueva contraseña.")
            return redirect(url_for('admin_cambiar_password'))
        if nueva == actual:
            flash("⚠️ La nueva contraseña debe ser distinta de la actual.")
            return redirect(url_for('admin_cambiar_password'))
        # Fortaleza mínima: mayúscula + minúscula + número
        tiene_mayus = any(c.isupper() for c in nueva)
        tiene_minus = any(c.islower() for c in nueva)
        tiene_num = any(c.isdigit() for c in nueva)
        if not (tiene_mayus and tiene_minus and tiene_num):
            flash("⚠️ La contraseña debe tener al menos una mayúscula, una minúscula y un número.")
            return redirect(url_for('admin_cambiar_password'))

        user.password_hash = generate_password_hash(nueva)
        user.must_change_password = False
        user.failed_attempts = 0
        user.locked_until = None
        db.session.commit()
        session.pop('forzar_cambio_password', None)
        try:
            registrar_auditoria('actualizar', 'Usuario', user.id,
                                valores_nuevos={'accion': 'cambio_password_propio'})
        except Exception:
            pass
        flash("✅ Contraseña actualizada correctamente.")
        return redirect(url_for('index'))
    return render_template('cambiar_password.html', usuario=user)




@app.route('/etiqueta/<int:item_id>')
@login_requerido
def etiqueta(item_id):
    """Imprime una etiqueta con código de barras + nombre + ubicación de un ítem.
    Esta ruta faltaba — el template inventario.html ya tenía un link a /etiqueta/<id>
    para el botón de impresión por fila, pero el endpoint no existía y daba 404."""
    item = Item.query.get_or_404(item_id)
    # Permitir solo a admin o al pañolero del área del ítem
    if (session.get('usuario_rol') != 'Admin'
            and item.especialidad_id != session.get('usuario_especialidad_id')):
        flash("❌ Sin permiso para ver esta etiqueta.")
        return redirect(url_for('ver_inventario'))
    return render_template('etiqueta.html', item=item)


@app.route('/etiquetas_lote')
@login_requerido
def etiquetas_lote():
    """Imprime TODAS las etiquetas del inventario de la especialidad actual,
    una por ítem, para corte rápido. Útil al inicio del año lectivo."""
    if session.get('usuario_rol') == 'Admin':
        esp_id = request.args.get('especialidad_id', type=int)
        if not esp_id:
            flash("⚠️ Como admin, indica ?especialidad_id=X.")
            return redirect(url_for('dashboard_admin'))
    else:
        esp_id = session.get('usuario_especialidad_id')
    if not esp_id:
        flash("❌ Sin especialidad activa.")
        return redirect(url_for('ver_inventario'))
    items = Item.query.filter_by(especialidad_id=esp_id).order_by(
        Item.categoria.asc(), Item.nombre.asc()).all()
    if not items:
        flash("⚠️ No hay ítems para imprimir.")
        return redirect(url_for('ver_inventario'))
    return render_template('etiquetas_lote.html', items=items)




@app.route('/inventario_listado')
@login_requerido
def inventario_listado():
    """Listado completo del inventario en formato tabla, imprimible (A4 horizontal)."""
    if session.get('usuario_rol') == 'Admin':
        esp_id = request.args.get('especialidad_id', type=int) or session.get('usuario_especialidad_id')
    else:
        esp_id = session.get('usuario_especialidad_id')
    if not esp_id:
        flash("❌ Sin especialidad activa.")
        return redirect(url_for('ver_inventario'))
    esp = Especialidad.query.get(esp_id)
    items = Item.query.filter_by(especialidad_id=esp_id).order_by(
        Item.categoria.asc(), Item.nombre.asc()).all()
    valor_total = sum((i.precio_unitario or 0) * (i.cantidad_total or 0) for i in items)
    return render_template('inventario_listado.html',
                           items=items,
                           especialidad_nombre=(esp.nombre if esp else 'Mi área'),
                           valor_total=valor_total,
                           fecha=datetime.now().strftime('%d-%m-%Y %H:%M'))


@app.route('/carnets_curso/<int:curso_id>')
@login_requerido
def carnets_curso(curso_id):
    """Página imprimible con los carnets (código de barras) de todos los alumnos
    de un curso. Genera el código de barras a los alumnos que aún no tengan."""
    curso = Curso.query.get_or_404(curso_id)
    if (session.get('usuario_rol') != 'Admin'
            and curso.especialidad_id != session.get('usuario_especialidad_id')):
        flash("❌ Sin permiso para ver los carnets de este curso.")
        return redirect(url_for('ver_inventario'))
    alumnos = Estudiante.query.filter_by(curso_id=curso.id, activo=True).order_by(
        Estudiante.numero_lista.asc(), Estudiante.nombre.asc()).all()
    # Asegurar que todos tengan código de barras
    generados = 0
    for a in alumnos:
        if not a.codigo_barras:
            a.codigo_barras = generar_codigo_barras_alumno()
            generados += 1
    if generados:
        db.session.commit()
    return render_template('carnets_curso.html', curso=curso, alumnos=alumnos)


if __name__ == '__main__':
    import threading, webbrowser, sys
    es_exe = getattr(sys, 'frozen', False)
    def abrir_navegador():
        import time; time.sleep(1.5)
        webbrowser.open('http://127.0.0.1:8080')
    PORT = 8080
    if es_exe:
        threading.Thread(target=abrir_navegador, daemon=True).start()
        app.run(host='127.0.0.1', port=PORT, debug=False, use_reloader=False)
    else:
        app.run(host='127.0.0.1', port=PORT, debug=True)
